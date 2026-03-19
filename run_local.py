#!/usr/bin/env python3
"""
DocLens — Standalone Local Runner
Runs without PostgreSQL/Redis using SQLite + in-memory rate limiting.
For production, use docker-compose.yml with full stack.
"""
import os
import sys
import uuid
import hashlib
import secrets
import time
import json
import logging
import re
import csv
from io import BytesIO, StringIO
from datetime import datetime, timezone, timedelta
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional
from collections import defaultdict

# ============================================================
# Configure logging
# ============================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger("doclens")

# ============================================================
# Try importing heavy deps, fallback gracefully
# ============================================================
try:
    import numpy as np
    import cv2
    HAS_CV2 = True
except ImportError:
    HAS_CV2 = False
    logger.warning("OpenCV not installed — image preprocessing will be basic")

HAS_EASYOCR = False
try:
    import easyocr
    HAS_EASYOCR = True
except ImportError:
    HAS_EASYOCR = False

HAS_TESSERACT = False
try:
    import pytesseract
    # Check if tesseract binary is available
    pytesseract.get_tesseract_version()
    HAS_TESSERACT = True
except Exception:
    HAS_TESSERACT = False

HAS_OPENPYXL = False
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

if not HAS_EASYOCR and not HAS_TESSERACT:
    logger.warning("No OCR engine available — will use demo mode")
elif HAS_EASYOCR:
    logger.info("Using EasyOCR (PRIMARY) with Tesseract (SECONDARY) — optimal for Russian text")
elif HAS_TESSERACT:
    logger.info("Using Tesseract OCR engine (EasyOCR not available)")

# ============================================================
# FastAPI Setup
# ============================================================
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Request, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
import uvicorn

# ============================================================
# In-Memory Database (SQLite-like)
# ============================================================
class InMemoryDB:
    """Simple in-memory store for local dev."""
    def __init__(self):
        self.tenants = {}
        self.api_keys = {}  # key_hash -> {tenant_id, name, ...}
        self.recognitions = {}
        self.usage = defaultdict(lambda: {"count": 0, "success": 0, "failed": 0, "ms": 0})
        # ML / Active Learning storage
        self.corrections = {}           # correction_id -> {recognition_id, field_name, original, corrected, ...}
        self.correction_patterns = {}   # pattern_id -> {doc_type, field_name, error, correction, count, confidence}
        self.training_runs = []         # [{id, status, accuracy_before, accuracy_after, ...}]
        self.ml_model_data = None       # In-memory model (substitution matrix + vocab)

    def create_tenant(self, name: str, email: str, plan: str = "free") -> dict:
        tid = str(uuid.uuid4())
        tenant = {
            "id": tid, "name": name, "email": email,
            "plan": plan, "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        self.tenants[tid] = tenant
        return tenant

    def create_api_key(self, tenant_id: str, name: str = "Default") -> tuple[dict, str]:
        raw_key = f"dl_live_{secrets.token_hex(24)}"
        key_hash = hashlib.sha256(raw_key.encode()).hexdigest()
        prefix = raw_key[8:16]
        kid = str(uuid.uuid4())
        record = {
            "id": kid, "tenant_id": tenant_id, "key_prefix": prefix,
            "key_hash": key_hash, "name": name, "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        self.api_keys[key_hash] = record
        return record, raw_key

    def auth_by_key(self, api_key: str) -> dict | None:
        key_hash = hashlib.sha256(api_key.encode()).hexdigest()
        record = self.api_keys.get(key_hash)
        if record and record["is_active"]:
            return self.tenants.get(record["tenant_id"])
        return None

    def save_recognition(self, tenant_id: str, result: dict) -> dict:
        rid = str(uuid.uuid4())
        rec = {
            "id": rid, "tenant_id": tenant_id,
            "created_at": datetime.now(timezone.utc).isoformat(),
            **result,
        }
        self.recognitions[rid] = rec

        # Update usage
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        key = f"{tenant_id}:{today}"
        self.usage[key]["count"] += 1
        if result.get("status") == "completed":
            self.usage[key]["success"] += 1
        else:
            self.usage[key]["failed"] += 1
        self.usage[key]["ms"] += result.get("processing_time_ms", 0)

        return rec

    def get_usage(self, tenant_id: str) -> dict:
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        key = f"{tenant_id}:{today}"
        u = self.usage[key]
        plan_limits = {"free": 100, "basic": 1000, "pro": 10000, "enterprise": 100000}
        tenant = self.tenants.get(tenant_id, {})
        limit = plan_limits.get(tenant.get("plan", "free"), 100)
        return {
            "plan": tenant.get("plan", "free"),
            "daily_limit": limit,
            "today_used": u["count"],
            "today_remaining": max(0, limit - u["count"]),
            "period_total": u["count"],
            "period_successful": u["success"],
            "period_failed": u["failed"],
            "avg_processing_ms": round(u["ms"] / max(u["count"], 1), 1),
        }

    # === ML / Active Learning Methods ===

    def save_correction(self, tenant_id: str, recognition_id: str,
                        document_type: str, field_name: str,
                        original_value: str, corrected_value: str,
                        confidence: float = 0.0) -> dict:
        """Save a user correction for ML learning."""
        if original_value.strip() == corrected_value.strip():
            return None

        cid = str(uuid.uuid4())
        # Auto-approve if we've seen this exact correction 3+ times
        matching = sum(
            1 for c in self.corrections.values()
            if c["document_type"] == document_type
            and c["field_name"] == field_name
            and c["original_value"] == original_value
            and c["corrected_value"] == corrected_value
            and c["status"] in ("approved", "trained")
        )
        status = "approved" if matching >= 2 else "pending"

        correction = {
            "id": cid, "tenant_id": tenant_id,
            "recognition_id": recognition_id,
            "document_type": document_type,
            "field_name": field_name,
            "original_value": original_value,
            "corrected_value": corrected_value,
            "original_confidence": confidence,
            "status": status,
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        self.corrections[cid] = correction

        # Auto-extract patterns when we accumulate enough corrections
        pending_count = sum(
            1 for c in self.corrections.values()
            if c["document_type"] == document_type
            and c["status"] in ("pending", "approved")
        )
        if pending_count >= 10:
            self._extract_patterns(document_type)

        return correction

    def _extract_patterns(self, document_type: str = None):
        """Extract error patterns from corrections."""
        from difflib import SequenceMatcher

        corrections = [
            c for c in self.corrections.values()
            if c["status"] in ("pending", "approved")
            and (document_type is None or c["document_type"] == document_type)
        ]

        # Group by (doc_type, field_name)
        grouped = defaultdict(list)
        for c in corrections:
            grouped[(c["document_type"], c["field_name"])].append(c)

        for (doc_type, field_name), field_corrections in grouped.items():
            # Character substitution patterns
            char_subs = defaultdict(int)
            for c in field_corrections:
                matcher = SequenceMatcher(None, c["original_value"], c["corrected_value"])
                for op, i1, i2, j1, j2 in matcher.get_opcodes():
                    if op == "replace":
                        old_chunk = c["original_value"][i1:i2]
                        new_chunk = c["corrected_value"][j1:j2]
                        if len(old_chunk) <= 3 and len(new_chunk) <= 3:
                            char_subs[(old_chunk, new_chunk)] += 1

            for (error, correct), count in char_subs.items():
                if count < 2:
                    continue
                # Upsert pattern
                existing = None
                for p in self.correction_patterns.values():
                    if (p["document_type"] == doc_type and p["field_name"] == field_name
                            and p["error_pattern"] == error and p["correction"] == correct):
                        existing = p
                        break

                if existing:
                    existing["occurrence_count"] += count
                    existing["confidence"] = min(0.99, 0.5 + existing["occurrence_count"] * 0.05)
                else:
                    pid = str(uuid.uuid4())
                    self.correction_patterns[pid] = {
                        "id": pid, "document_type": doc_type, "field_name": field_name,
                        "error_pattern": error, "correction": correct,
                        "pattern_type": "char_substitution",
                        "occurrence_count": count,
                        "confidence": min(0.99, 0.5 + count * 0.05),
                        "is_active": True,
                    }

    def get_correction_stats(self, document_type: str = None) -> dict:
        """Get correction and pattern stats."""
        corrections = [c for c in self.corrections.values()
                       if document_type is None or c["document_type"] == document_type]
        by_status = defaultdict(int)
        for c in corrections:
            by_status[c["status"]] += 1

        patterns = [p for p in self.correction_patterns.values()
                    if p["is_active"]
                    and (document_type is None or p["document_type"] == document_type)]

        return {
            "total_corrections": len(corrections),
            "by_status": dict(by_status),
            "active_patterns": len(patterns),
            "ready_for_training": len(corrections) >= 20,
        }


db = InMemoryDB()


# ============================================================
# ML Field Corrector (inline for standalone mode)
# ============================================================

# Common OCR char fixes: Latin/digit ↔ Cyrillic
_CYRILLIC_OCR_FIXES = {
    "A": "А", "B": "В", "C": "С", "E": "Е", "H": "Н",
    "K": "К", "M": "М", "O": "О", "P": "Р", "T": "Т",
    "X": "Х", "a": "а", "c": "с", "e": "е", "o": "о",
    "p": "р", "x": "х", "y": "у",
}
_DIGIT_FIXES = {
    "О": "0", "о": "0", "З": "3", "з": "3",
    "б": "6", "В": "8", "l": "1", "I": "1", "S": "5",
}

# Field format rules for Russian documents
_FIELD_RULES = {
    "passport_rf": {
        "series": {"digits": True, "format": r"^\d{2}\s\d{2}$"},
        "number": {"digits": True, "format": r"^\d{6}$"},
        "department_code": {"digits": True, "format": r"^\d{3}-\d{3}$"},
        "issue_date": {"date": True},
        "birth_date": {"date": True},
        "sex": {"pattern": r"^[МЖ]$"},
        "last_name": {"cyrillic": True},
        "first_name": {"cyrillic": True},
        "patronymic": {"cyrillic": True},
    },
    "snils": {
        "number": {"digits": True, "format": r"^\d{3}-\d{3}-\d{3}\s\d{2}$"},
    },
}


def ml_correct_fields(document_type: str, fields: dict) -> dict:
    """Apply ML corrections to OCR-extracted fields.

    3-tier correction:
    1. Format fixes (domain knowledge)
    2. Character substitution fixes (Cyrillic ↔ Latin/digits)
    3. Learned patterns from user feedback
    """
    rules = _FIELD_RULES.get(document_type, {})
    corrected = {}

    for name, fdata in fields.items():
        if not isinstance(fdata, dict) or "value" not in fdata:
            corrected[name] = fdata
            continue

        value = fdata["value"]
        original = value
        field_rules = rules.get(name, {})

        # Tier 1: Format corrections
        if field_rules.get("digits"):
            # Replace Cyrillic lookalikes with digits
            value = "".join(_DIGIT_FIXES.get(ch, ch) for ch in value)

            # Fix department code format
            if name == "department_code":
                digits = re.sub(r"\D", "", value)
                if len(digits) == 6:
                    value = f"{digits[:3]}-{digits[3:]}"

            # Fix SNILS format
            if name == "number" and document_type == "snils":
                digits = re.sub(r"\D", "", value)
                if len(digits) == 11:
                    value = f"{digits[:3]}-{digits[3:6]}-{digits[6:9]} {digits[9:]}"

        if field_rules.get("date"):
            value = re.sub(r"[\-/]", ".", value)
            m = re.match(r"(\d{1,2})\.(\d{1,2})\.(\d{4})", value)
            if m:
                value = f"{m.group(1).zfill(2)}.{m.group(2).zfill(2)}.{m.group(3)}"

        if field_rules.get("cyrillic"):
            # Replace Latin lookalikes with Cyrillic
            value = "".join(
                _CYRILLIC_OCR_FIXES.get(ch, ch) if not ch.isdigit() else ch
                for ch in value
            )
            # Proper capitalization: ИВАНОВ → Иванов
            if value == value.upper() and len(value) > 1:
                parts = value.split("-")
                value = "-".join(p.capitalize() for p in parts)

        # Tier 2: Learned patterns from user feedback
        for p in db.correction_patterns.values():
            if (p["is_active"] and p["confidence"] >= 0.6
                    and p["document_type"] == document_type
                    and p["field_name"] == name):
                if p["pattern_type"] == "char_substitution" and p["error_pattern"] in value:
                    value = value.replace(p["error_pattern"], p["correction"])

        # Tier 3: ML model correction (if trained)
        if db.ml_model_data and fdata.get("confidence", 1.0) < 0.9:
            try:
                sub_matrix = db.ml_model_data.get("substitution_matrix")
                vocab = db.ml_model_data.get("vocab")
                if sub_matrix is not None and vocab is not None:
                    new_chars = []
                    for ch in value:
                        if ch in vocab:
                            idx = vocab[ch]
                            if idx < sub_matrix.shape[0]:
                                row = sub_matrix[idx]
                                best_idx = int(np.argmax(row))
                                if row[best_idx] > 0.7 and row[idx] < 0.5 and best_idx != idx:
                                    rev = {v: k for k, v in vocab.items()}
                                    if best_idx in rev:
                                        new_chars.append(rev[best_idx])
                                        continue
                        new_chars.append(ch)
                    value = "".join(new_chars)
            except Exception as e:
                logger.warning(f"ML model correction failed: {e}")

        # Build result
        cf = dict(fdata)
        if value != original:
            cf["value"] = value
            cf["original_ocr_value"] = original
            cf["ml_corrected"] = True
            cf["confidence"] = min(0.99, fdata.get("confidence", 0.5) + 0.05)
            logger.info(f"ML corrected {name}: '{original}' → '{value}'")
        corrected[name] = cf

    return corrected


def ml_train_model(document_type: str = None, force: bool = False) -> dict:
    """Train ML model from accumulated corrections.

    Builds a character substitution probability matrix.
    """
    corrections = [
        c for c in db.corrections.values()
        if c["status"] in ("pending", "approved")
        and (document_type is None or c["document_type"] == document_type)
    ]

    if len(corrections) < 20 and not force:
        return {
            "status": "failed",
            "error": f"Need at least 20 corrections, have {len(corrections)}",
        }

    start = time.time()

    # Extract patterns first
    db._extract_patterns(document_type)

    # Build substitution matrix
    all_chars = set()
    for c in corrections:
        all_chars.update(c["original_value"])
        all_chars.update(c["corrected_value"])

    vocab = {ch: i for i, ch in enumerate(sorted(all_chars))}
    n = len(vocab)
    if n == 0:
        return {"status": "failed", "error": "No character data"}

    counts = np.zeros((n, n), dtype=np.float64)
    for c in corrections:
        orig, corr = c["original_value"], c["corrected_value"]
        for o_ch, c_ch in zip(orig, corr) if len(orig) == len(corr) else []:
            if o_ch in vocab and c_ch in vocab:
                counts[vocab[o_ch]][vocab[c_ch]] += 1

    row_sums = counts.sum(axis=1, keepdims=True)
    row_sums[row_sums == 0] = 1
    matrix = counts / row_sums
    matrix = matrix * 0.7 + np.eye(n) * 0.3
    row_sums = matrix.sum(axis=1, keepdims=True)
    matrix = matrix / row_sums

    # Evaluate
    test_n = max(1, len(corrections) // 5)
    test = corrections[-test_n:]
    acc_before = sum(1 for c in test if c["original_value"] == c["corrected_value"]) / len(test)

    correct = 0
    for c in test:
        predicted = "".join(
            (lambda idx, row: (
                {v: k for k, v in vocab.items()}.get(int(np.argmax(row)), ch)
                if row[int(np.argmax(row))] > 0.6 and int(np.argmax(row)) != idx
                else ch
            ))(vocab.get(ch, 0), matrix[vocab[ch]] if ch in vocab and vocab[ch] < n else np.zeros(n))
            for ch in c["original_value"]
        )
        if predicted == c["corrected_value"]:
            correct += 1
    acc_after = correct / len(test)

    # Save model in memory
    db.ml_model_data = {"substitution_matrix": matrix, "vocab": vocab}

    # Mark corrections as trained
    for c in corrections:
        c["status"] = "trained"

    duration = int(time.time() - start)
    run_id = str(uuid.uuid4())
    run = {
        "id": run_id, "status": "completed",
        "corrections_count": len(corrections),
        "patterns_generated": len(db.correction_patterns),
        "accuracy_before": round(acc_before, 4),
        "accuracy_after": round(acc_after, 4),
        "duration_seconds": duration,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    db.training_runs.append(run)
    logger.info(f"ML training complete: {acc_before:.1%} → {acc_after:.1%}, "
                f"{len(corrections)} corrections, {duration}s")
    return run


# ============================================================
# OCR Pipeline (lightweight version)
# ============================================================

@dataclass
class OCRLine:
    text: str
    confidence: float
    bbox: list = field(default_factory=list)
    center_x: float = 0.0
    center_y: float = 0.0

@dataclass
class FieldResult:
    value: str
    confidence: float
    source: str = "ocr"
    auto_fill: bool = False
    needs_review: bool = False


class LightweightPipeline:
    """OCR pipeline using EasyOCR (primary) and Tesseract (secondary)."""

    def __init__(self):
        self.easyocr_reader = None
        self.easyocr_reader_en = None
        self.engine = "none"
        self.tessdata_quality = "unknown"

        # Initialize EasyOCR as PRIMARY engine for Russian language support
        # EasyOCR handles soft signs (ДЬ) much better than Tesseract
        if HAS_EASYOCR:
            try:
                logger.info("Initializing EasyOCR for Russian language support (PRIMARY)...")
                self.easyocr_reader = easyocr.Reader(['ru', 'en'], gpu=False)
                self.engine = "easyocr"
                logger.info("EasyOCR reader initialized successfully")
                # Initialize English-only reader for MRZ (Machine Readable Zone)
                # to avoid Cyrillic misrecognition of Latin characters
                try:
                    logger.info("Initializing English-only EasyOCR reader for MRZ...")
                    self.easyocr_reader_en = easyocr.Reader(['en'], gpu=False)
                    logger.info("English-only EasyOCR reader for MRZ initialized successfully")
                except Exception as e_en:
                    logger.warning(f"English-only EasyOCR reader initialization failed: {e_en}")
                    self.easyocr_reader_en = None
            except Exception as e:
                logger.error(f"EasyOCR initialization failed: {e}")
                self.easyocr_reader_en = None
                # Fall back to Tesseract
                if HAS_TESSERACT:
                    self.engine = "tesseract"
                    logger.info("Tesseract OCR initialized as fallback")
                    quality, size_mb, path = self._check_tessdata_quality()
                    self.tessdata_quality = quality
        elif HAS_TESSERACT:
            self.engine = "tesseract"
            self.easyocr_reader_en = None
            logger.info("Tesseract OCR initialized successfully")
            # Check tessdata quality
            quality, size_mb, path = self._check_tessdata_quality()
            self.tessdata_quality = quality

    _cached_tess_lang = None

    def _detect_tess_lang(self):
        """Detect available Tesseract languages (cached)."""
        if LightweightPipeline._cached_tess_lang is not None:
            return LightweightPipeline._cached_tess_lang
        try:
            available = pytesseract.get_languages()
            if "rus" in available:
                LightweightPipeline._cached_tess_lang = "rus+eng"
            else:
                logger.warning("Tesseract: Russian language not found, using English only. Install: brew install tesseract-lang")
                LightweightPipeline._cached_tess_lang = "eng"
        except Exception:
            LightweightPipeline._cached_tess_lang = "eng"
        return LightweightPipeline._cached_tess_lang

    def _check_tessdata_quality(self):
        """Check if tessdata_best is being used (better for Russian OCR)."""
        try:
            import subprocess
            # Find tessdata path
            result = subprocess.run(["tesseract", "--print-parameters"], capture_output=True, text=True, timeout=5)
            # Check for tessdata path in output
            tessdata_path = None
            for line in result.stderr.split("\n") + result.stdout.split("\n"):
                if "tessdata" in line.lower():
                    parts = line.split()
                    for p in parts:
                        if "tessdata" in p.lower() and "/" in p:
                            tessdata_path = p
                            break

            if not tessdata_path:
                # Try common paths
                import glob
                for path in ["/opt/homebrew/share/tessdata", "/usr/local/share/tessdata", "/usr/share/tesseract-ocr/5/tessdata"]:
                    if os.path.exists(path):
                        tessdata_path = path
                        break

            if tessdata_path:
                rus_file = os.path.join(tessdata_path, "rus.traineddata")
                if os.path.exists(rus_file):
                    size_mb = os.path.getsize(rus_file) / (1024 * 1024)
                    # tessdata_fast: ~2-4MB, tessdata: ~15-20MB, tessdata_best: ~30-40MB
                    if size_mb < 5:
                        quality = "fast"
                        logger.warning(f"Tesseract using FAST tessdata ({size_mb:.1f}MB) — poor for Russian names!")
                        logger.warning(f"For better OCR, download tessdata_best: curl -L -o {rus_file} https://github.com/tesseract-ocr/tessdata_best/raw/main/rus.traineddata")
                    elif size_mb < 25:
                        quality = "standard"
                        logger.info(f"Tesseract using standard tessdata ({size_mb:.1f}MB)")
                    else:
                        quality = "best"
                        logger.info(f"Tesseract using BEST tessdata ({size_mb:.1f}MB) — optimal for Russian!")
                    return quality, size_mb, tessdata_path
        except Exception as e:
            logger.debug(f"tessdata quality check failed: {e}")
        return "unknown", 0, ""

    def _tess_image_to_lines(self, pil_img, lang, config="") -> list:
        """Run Tesseract on a PIL image and return OCRLine list."""
        from PIL import Image as PILImage
        lines = []
        data = pytesseract.image_to_data(pil_img, lang=lang, config=config, output_type=pytesseract.Output.DICT)
        n = len(data["text"])
        line_groups = {}
        for i in range(n):
            text = data["text"][i].strip()
            conf = int(data["conf"][i])
            if not text or conf < 10:
                continue
            key = (data["block_num"][i], data["par_num"][i], data["line_num"][i])
            if key not in line_groups:
                line_groups[key] = {"texts": [], "confs": [], "left": 9999, "top": 9999, "right": 0, "bottom": 0}
            g = line_groups[key]
            g["texts"].append(text)
            g["confs"].append(conf)
            x, y, w, h = data["left"][i], data["top"][i], data["width"][i], data["height"][i]
            g["left"] = min(g["left"], x)
            g["top"] = min(g["top"], y)
            g["right"] = max(g["right"], x + w)
            g["bottom"] = max(g["bottom"], y + h)

        for key in sorted(line_groups.keys()):
            g = line_groups[key]
            line_text = " ".join(g["texts"])
            avg_conf = sum(g["confs"]) / len(g["confs"]) / 100.0
            bbox = [[g["left"], g["top"]], [g["right"], g["top"]],
                    [g["right"], g["bottom"]], [g["left"], g["bottom"]]]
            cx = (g["left"] + g["right"]) / 2
            cy = (g["top"] + g["bottom"]) / 2
            lines.append(OCRLine(text=line_text, confidence=avg_conf, bbox=bbox,
                                 center_x=cx, center_y=cy))
        return lines

    def _ocr_tesseract(self, image) -> list:
        """Run Tesseract OCR on an image, return list of OCRLine."""
        from PIL import Image as PILImage
        rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        pil_img = PILImage.fromarray(rgb)
        tess_lang = self._detect_tess_lang()
        lines = self._tess_image_to_lines(pil_img, tess_lang)
        return lines

    def _ocr_tesseract_multi_strategy(self, image) -> list:
        """Run Tesseract with multiple preprocessing strategies, merge unique lines."""
        from PIL import Image as PILImage
        tess_lang = self._detect_tess_lang()
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if len(image.shape) == 3 else image

        all_strategy_lines = []

        # Strategy 1: CLAHE on brightness-corrected image
        try:
            clahe1 = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8, 8))
            enhanced1 = clahe1.apply(gray)
            pil1 = PILImage.fromarray(enhanced1)
            lines1 = self._tess_image_to_lines(pil1, tess_lang)
            cyr1 = sum(1 for l in lines1 for c in l.text if '\u0400' <= c <= '\u04FF')
            all_strategy_lines.append(("clahe", lines1, cyr1))
            logger.info(f"Multi-OCR CLAHE: {len(lines1)} lines, {cyr1} Cyrillic chars")
        except Exception as e:
            logger.warning(f"Multi-OCR CLAHE failed: {e}")

        # Strategy 2: Adaptive threshold
        try:
            thresh2 = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 31, 10)
            pil2 = PILImage.fromarray(thresh2)
            lines2 = self._tess_image_to_lines(pil2, tess_lang)
            cyr2 = sum(1 for l in lines2 for c in l.text if '\u0400' <= c <= '\u04FF')
            all_strategy_lines.append(("adaptive", lines2, cyr2))
            logger.info(f"Multi-OCR Adaptive: {len(lines2)} lines, {cyr2} Cyrillic chars")
        except Exception as e:
            logger.warning(f"Multi-OCR Adaptive failed: {e}")

        # Strategy 3: Otsu binarization
        try:
            _, thresh3 = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            pil3 = PILImage.fromarray(thresh3)
            lines3 = self._tess_image_to_lines(pil3, tess_lang)
            cyr3 = sum(1 for l in lines3 for c in l.text if '\u0400' <= c <= '\u04FF')
            all_strategy_lines.append(("otsu", lines3, cyr3))
            logger.info(f"Multi-OCR Otsu: {len(lines3)} lines, {cyr3} Cyrillic chars")
        except Exception as e:
            logger.warning(f"Multi-OCR Otsu failed: {e}")

        # Strategy 4: Sharpen + aggressive CLAHE
        try:
            kernel = np.array([[-1, -1, -1], [-1, 9, -1], [-1, -1, -1]])
            sharpened = cv2.filter2D(gray, -1, kernel)
            clahe4 = cv2.createCLAHE(clipLimit=4.0, tileGridSize=(8, 8))
            enhanced4 = clahe4.apply(sharpened)
            pil4 = PILImage.fromarray(enhanced4)
            lines4 = self._tess_image_to_lines(pil4, tess_lang)
            cyr4 = sum(1 for l in lines4 for c in l.text if '\u0400' <= c <= '\u04FF')
            all_strategy_lines.append(("sharpen_clahe", lines4, cyr4))
            logger.info(f"Multi-OCR Sharpen+CLAHE: {len(lines4)} lines, {cyr4} Cyrillic chars")
        except Exception as e:
            logger.warning(f"Multi-OCR Sharpen+CLAHE failed: {e}")

        if not all_strategy_lines:
            return []

        # Merge: start with the strategy that has most Cyrillic chars,
        # then add unique lines from other strategies
        all_strategy_lines.sort(key=lambda x: -x[2])
        best_name, best_lines, best_cyr = all_strategy_lines[0]
        logger.info(f"Multi-OCR primary: {best_name} ({best_cyr} Cyrillic, {len(best_lines)} lines)")

        merged = list(best_lines)
        existing_texts = {l.text.strip().upper() for l in merged}

        for name, lines, cyr in all_strategy_lines[1:]:
            for line in lines:
                t = line.text.strip().upper()
                # Skip if we already have very similar text
                if t in existing_texts:
                    continue
                # Skip very short or garbage text
                if len(t) < 2:
                    continue
                # Check if this line contains Cyrillic text not in existing lines
                line_cyr = sum(1 for c in t if '\u0400' <= c <= '\u04FF')
                if line_cyr < 2:
                    continue
                # Check if overlaps spatially with existing line (within 20px vertical)
                has_overlap = False
                for existing in merged:
                    if abs(line.center_y - existing.center_y) < 20 and abs(line.center_x - existing.center_x) < 50:
                        # Same spatial region — keep the one with more Cyrillic chars
                        existing_cyr = sum(1 for c in existing.text if '\u0400' <= c <= '\u04FF')
                        if line_cyr > existing_cyr:
                            # Replace with better version
                            idx = merged.index(existing)
                            merged[idx] = line
                            existing_texts.discard(existing.text.strip().upper())
                            existing_texts.add(t)
                            logger.info(f"Multi-OCR merge: replaced '{existing.text}' with '{line.text}' from {name}")
                        has_overlap = True
                        break
                if not has_overlap:
                    merged.append(line)
                    existing_texts.add(t)
                    logger.info(f"Multi-OCR merge: added '{line.text}' from {name}")

        logger.info(f"Multi-OCR merged: {len(merged)} total lines")
        return merged

    def _ocr_easyocr(self, image) -> list:
        """Run EasyOCR on image and return OCRLine list.
        EasyOCR has better support for Russian text, especially soft signs (ДЬ).
        """
        if not self.easyocr_reader:
            logger.warning("EasyOCR reader not initialized")
            return []

        try:
            # Resize image if too large (memory constraint)
            h_orig, w_orig = image.shape[:2]
            max_dim = 1600  # Good balance of quality vs memory on Mac (8GB+)
            coord_scale = 1.0
            if max(h_orig, w_orig) > max_dim:
                coord_scale = max(h_orig, w_orig) / max_dim
                scale = max_dim / max(h_orig, w_orig)
                image = cv2.resize(image, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
                logger.info(f"EasyOCR: resized from {w_orig}x{h_orig} to {image.shape[1]}x{image.shape[0]} (coord_scale={coord_scale:.2f})")

            # EasyOCR expects BGR or RGB
            results = self.easyocr_reader.readtext(image, detail=1)
            lines = []

            if not results:
                logger.info("EasyOCR returned no results")
                return lines

            # Log raw EasyOCR output for debugging
            for (bbox, text, conf) in results[:20]:
                logger.info(f"EasyOCR raw: '{text}' conf={conf:.2f}")

            # Group results by approximate vertical position (within ~15 pixels)
            line_groups = {}
            for (bbox, text, confidence) in results:
                if not bbox or not text:
                    continue

                # Calculate bounding box from polygon points, scale back to original coords
                xs = [p[0] * coord_scale for p in bbox]
                ys = [p[1] * coord_scale for p in bbox]
                left = min(xs)
                top = min(ys)
                right = max(xs)
                bottom = max(ys)
                center_x = (left + right) / 2
                center_y = (top + bottom) / 2

                # Group by approximate vertical position (line_y = round to nearest 15px * coord_scale)
                group_tolerance = max(15, int(15 * coord_scale))
                line_y = round(center_y / group_tolerance) * group_tolerance
                if line_y not in line_groups:
                    line_groups[line_y] = {
                        "texts": [],
                        "confidences": [],
                        "left": left,
                        "top": top,
                        "right": right,
                        "bottom": bottom,
                        "min_x": left,
                    }
                g = line_groups[line_y]
                g["texts"].append(text)
                g["confidences"].append(confidence)
                g["left"] = min(g["left"], left)
                g["top"] = min(g["top"], top)
                g["right"] = max(g["right"], right)
                g["bottom"] = max(g["bottom"], bottom)

            # Convert groups to OCRLine objects
            for line_y in sorted(line_groups.keys()):
                g = line_groups[line_y]
                line_text = " ".join(g["texts"])
                avg_conf = sum(g["confidences"]) / len(g["confidences"])
                bbox_list = [[g["left"], g["top"]], [g["right"], g["top"]],
                            [g["right"], g["bottom"]], [g["left"], g["bottom"]]]
                cx = (g["left"] + g["right"]) / 2
                cy = (g["top"] + g["bottom"]) / 2
                lines.append(OCRLine(text=line_text, confidence=avg_conf, bbox=bbox_list,
                                   center_x=cx, center_y=cy))

            cyr_count = sum(1 for l in lines for c in l.text if '\u0400' <= c <= '\u04FF')
            logger.info(f"EasyOCR extracted {len(lines)} lines, {cyr_count} Cyrillic chars")
            return lines

        except Exception as e:
            logger.error(f"EasyOCR failed: {e}")
            return []

    def _ocr_ensemble(self, image, image_original=None) -> list:
        """Run both EasyOCR and Tesseract, merge results intelligently.
        EasyOCR is PRIMARY (used as base) because it handles Russian soft signs (ДЬ) much better.
        Tesseract is SECONDARY (supplemental) to add lines not covered by EasyOCR.
        For overlapping regions, prefer the line with more Cyrillic characters.

        image: preprocessed image for Tesseract
        image_original: original/lightly processed image for EasyOCR (EasyOCR has its own preprocessing)
        """
        # EasyOCR works best with clean, minimally-processed images
        easyocr_input = image_original if image_original is not None else image
        easyocr_lines = self._ocr_easyocr(easyocr_input)
        tesseract_lines = self._ocr_tesseract_multi_strategy(image)

        if not easyocr_lines:
            logger.info("Ensemble: EasyOCR returned no lines, using Tesseract only")
            return tesseract_lines
        if not tesseract_lines:
            logger.info("Ensemble: Tesseract returned no lines, using EasyOCR only")
            return easyocr_lines

        # === START WITH EASYOCR AS PRIMARY ===
        merged = list(easyocr_lines)
        logger.info(f"Ensemble starting with {len(easyocr_lines)} EasyOCR lines (PRIMARY) and {len(tesseract_lines)} Tesseract lines (SECONDARY)")

        for tess_line in tesseract_lines:
            # Check if this Tesseract line overlaps spatially with any EasyOCR line
            best_overlap_idx = None
            best_distance = 9999
            min_overlap_threshold = 20  # pixels

            for i, easy_line in enumerate(merged):
                # Check spatial overlap (within ~20px vertical, ~50px horizontal)
                vert_dist = abs(tess_line.center_y - easy_line.center_y)
                horiz_dist = abs(tess_line.center_x - easy_line.center_x)

                if vert_dist < min_overlap_threshold and horiz_dist < 50:
                    distance = vert_dist + horiz_dist * 0.3
                    if distance < best_distance:
                        best_distance = distance
                        best_overlap_idx = i

            if best_overlap_idx is not None:
                # There's an overlap — compare quality: prefer the one with MORE Cyrillic
                easy_line = merged[best_overlap_idx]
                easy_cyr = sum(1 for c in easy_line.text if '\u0400' <= c <= '\u04FF')
                tess_cyr = sum(1 for c in tess_line.text if '\u0400' <= c <= '\u04FF')

                # EasyOCR wins ties (because it reads ДЬ correctly)
                if tess_cyr > easy_cyr:
                    logger.info(f"Ensemble: replacing EasyOCR '{easy_line.text}' "
                              f"(cyr={easy_cyr}) with Tesseract '{tess_line.text}' (cyr={tess_cyr})")
                    merged[best_overlap_idx] = tess_line
                else:
                    logger.debug(f"Ensemble: keeping EasyOCR '{easy_line.text}' "
                               f"(cyr={easy_cyr}) over Tesseract '{tess_line.text}' (cyr={tess_cyr})")
            else:
                # No overlap — add the unique Tesseract line only if it has Cyrillic
                tess_cyr = sum(1 for c in tess_line.text if '\u0400' <= c <= '\u04FF')
                if tess_cyr > 0:  # Only add if it has Cyrillic
                    merged.append(tess_line)
                    logger.info(f"Ensemble: added unique Tesseract line '{tess_line.text}'")

        logger.info(f"Ensemble merged: {len(merged)} total lines")
        return merged

    @staticmethod
    def _normalize_mrz_k_to_filler(line: str) -> str:
        """
        Tesseract often reads MRZ '<' filler as 'K'.
        Conservative replacement to avoid damaging real K in names (DIAKOV, KOZLOV, etc.):
        - 2+ consecutive K's → replace all with <
        - K at position 1 (P< prefix) → <
        - Trailing K's → <
        - K adjacent to already-converted < → <
        """
        chars = list(line)
        n = len(chars)

        # Pass 1: Replace runs of 2+ consecutive K's with <
        i = 0
        while i < n:
            if chars[i] == "K":
                j = i
                while j < n and chars[j] == "K":
                    j += 1
                if j - i >= 2:
                    for idx in range(i, j):
                        chars[idx] = "<"
                i = j
            else:
                i += 1

        # Pass 2: For line starting with P, position 1 should be <
        if n >= 5 and chars[0] == "P" and chars[1] == "K":
            chars[1] = "<"

        # Pass 3: Trailing K's (3+ at end) are filler
        trail = 0
        for i in range(n - 1, max(n - 20, -1), -1):
            if chars[i] in ("K", "<"):
                trail += 1
            else:
                break
        if trail >= 3:
            for i in range(n - trail, n):
                chars[i] = "<"

        # Pass 4: K immediately adjacent to < is also likely < (propagate)
        changed = True
        iterations = 0
        while changed and iterations < 5:
            changed = False
            iterations += 1
            for i in range(n):
                if chars[i] == "K":
                    adj_filler = False
                    if i > 0 and chars[i-1] == "<":
                        adj_filler = True
                    if i < n-1 and chars[i+1] == "<":
                        adj_filler = True
                    if adj_filler:
                        chars[i] = "<"
                        changed = True

        return "".join(chars)

    def _ocr_mrz_zone(self, image) -> list:
        """Dedicated MRZ OCR: crop bottom 35% of image, heavy binarization, eng-only OCR."""
        from PIL import Image as PILImage
        h, w = image.shape[:2]
        # MRZ is at the bottom of the passport — crop bottom 35%
        mrz_crop = image[int(h * 0.60):, :]
        # Convert to grayscale and apply aggressive binarization
        gray = cv2.cvtColor(mrz_crop, cv2.COLOR_BGR2GRAY)
        # Resize up for better OCR on small text
        scale = max(1.0, 1500 / max(gray.shape[1], 1))
        if scale > 1.0:
            gray = cv2.resize(gray, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
        # Otsu binarization — best for MRZ
        _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        # Also try inverted
        _, binary_inv = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)

        mrz_lines = []
        for bimg in [binary, binary_inv]:
            pil_img = PILImage.fromarray(bimg)
            # Use eng only + PSM 6 (single block of text) for MRZ
            try:
                raw_text = pytesseract.image_to_string(pil_img, lang="eng", config="--psm 6 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789<")
                logger.info(f"MRZ zone OCR raw:\n{raw_text[:300]}")
                for line in raw_text.split("\n"):
                    clean = line.strip().replace(" ", "")
                    if len(clean) >= 28:
                        # Step 1: Basic normalization
                        normalized = ""
                        for c in clean.upper():
                            if c.isalpha() or c.isdigit() or c == "<":
                                normalized += c
                            else:
                                normalized += "<"
                        # Step 2: Smart K → < replacement
                        normalized = self._normalize_mrz_k_to_filler(normalized)
                        if re.match(r"^[A-Z0-9<]+$", normalized):
                            mrz_lines.append(normalized)
                            logger.info(f"MRZ zone candidate (normalized): {normalized}")
            except Exception as e:
                logger.warning(f"MRZ zone OCR failed: {e}")
            if len(mrz_lines) >= 2:
                break
        return mrz_lines

    def _correct_ocr_names(self, text: str) -> str:
        """Post-process OCR text to fix common Tesseract mistakes with Russian names.
        Tesseract often misreads ДЬ as ЛЪ, ДБ, ЛЬ, ТЬ, etc."""
        if not text or len(text) < 2:
            return text
        original = text
        # Common Tesseract misreadings for Russian characters
        corrections = [
            # ДЬ problems (soft sign combinations)
            ("ЛЪ", "ДЬ"),
            ("ЛЬ", "ДЬ"),  # only at word start
            ("ТЬ", "ДЬ"),  # at word start only
            ("ДБ", "ДЬ"),
            ("ЛБ", "ДЬ"),
            # Individual character fixes
            ("Ъ", "Ь"),  # hard sign → soft sign (more common in names)
        ]
        # Apply corrections at the start of words (most name-relevant position)
        for wrong, right in corrections:
            if text.upper().startswith(wrong):
                text = right + text[len(wrong):]
        # Also fix inside words but more conservatively
        # ЛЪЯ → ДЬЯ (very specific pattern for ДЬЯКОВ, ДЬЯКОНОВ, etc.)
        text = re.sub(r"ЛЪЯ", "ДЬЯ", text)
        text = re.sub(r"ЛъЯ", "ДьЯ", text)
        text = re.sub(r"лъя", "дья", text)
        # Also handle case where the Д itself is misread
        # ПЬЯКОВ → ДЬЯКОВ (П→Д at start of surname before ЬЯ)
        if re.match(r"^[ПпЛлТт][ЬьЪъ][ЯяЕеИиОоУу]", text):
            text = "Д" + text[1:]
            if text[1] == "Ъ" or text[1] == "ъ":
                text = text[0] + "Ь" + text[2:]
        if text != original:
            logger.info(f"OCR correction: '{original}' → '{text}'")
        return text

    def _ocr_name_zone(self, image) -> list:
        """Dedicated name-zone OCR for internal Russian passports.
        Crops the right-center area where FIO is printed and tries
        multiple aggressive preprocessing + PSM modes."""
        from PIL import Image as PILImage
        h, w = image.shape[:2]
        tess_lang = self._detect_tess_lang()

        # On internal passport (photo page), names are roughly:
        # - Right 65% of image (left side is the photo or registration page)
        # - Vertical: 25%-75% of image height (wider range to catch all names)
        # For a two-page spread, names are in right half, center vertically
        crop_x1 = int(w * 0.35)
        crop_y1 = int(h * 0.25)
        crop_y2 = int(h * 0.75)
        name_crop = image[crop_y1:crop_y2, crop_x1:]
        crop_h, crop_w = name_crop.shape[:2]
        logger.info(f"Name-zone crop: x={crop_x1}-{w}, y={crop_y1}-{crop_y2}, size={crop_w}x{crop_h}")

        gray = cv2.cvtColor(name_crop, cv2.COLOR_BGR2GRAY)

        # Scale up more aggressively for better OCR
        scale = max(1.0, 1800 / max(crop_w, 1))
        if scale > 1.0:
            gray = cv2.resize(gray, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
        logger.info(f"Name-zone scaled: {gray.shape[1]}x{gray.shape[0]} (scale={scale:.2f})")

        all_lines = []

        # ---- Preprocessing variants ----
        preprocessings = []

        # 1. CLAHE aggressive
        clahe = cv2.createCLAHE(clipLimit=4.0, tileGridSize=(6, 6))
        preprocessings.append(("clahe4", clahe.apply(gray)))

        # 2. Sharpen + CLAHE
        kernel = np.array([[-1, -1, -1], [-1, 9, -1], [-1, -1, -1]])
        sharpened = cv2.filter2D(gray, -1, kernel)
        clahe2 = cv2.createCLAHE(clipLimit=5.0, tileGridSize=(8, 8))
        preprocessings.append(("sharpen5", clahe2.apply(sharpened)))

        # 3. Bilateral filter + CLAHE (reduces noise while preserving edges)
        bilateral = cv2.bilateralFilter(gray, 9, 75, 75)
        clahe3 = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8, 8))
        preprocessings.append(("bilateral", clahe3.apply(bilateral)))

        # 4. Adaptive threshold
        preprocessings.append(("adaptive", cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 21, 8)))

        # 5. Otsu
        _, otsu = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        preprocessings.append(("otsu", otsu))

        # 6. Morphological: dilate then CLAHE (thickens thin strokes — helps with ДЬ)
        kernel_dilate = np.ones((2, 2), np.uint8)
        dilated = cv2.dilate(gray, kernel_dilate, iterations=1)
        clahe_d = cv2.createCLAHE(clipLimit=4.0, tileGridSize=(6, 6))
        preprocessings.append(("dilate_clahe", clahe_d.apply(dilated)))

        # 7. Heavy contrast + sharpen (for dark/low-contrast photos)
        alpha_heavy = 2.0
        beta_heavy = 30
        contrast_heavy = cv2.convertScaleAbs(gray, alpha=alpha_heavy, beta=beta_heavy)
        sharp_heavy = cv2.filter2D(contrast_heavy, -1, kernel)
        preprocessings.append(("heavy_contrast", sharp_heavy))

        # 8. Inverted threshold (white text on dark — sometimes passport OCR reads better)
        _, inv_otsu = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
        preprocessings.append(("inv_otsu", inv_otsu))

        for name, processed in preprocessings:
            for psm in [6, 4, 3]:  # PSM 6=block, 4=column of text, 3=auto
                try:
                    pil_img = PILImage.fromarray(processed)
                    config = f"--psm {psm}"
                    lines = self._tess_image_to_lines(pil_img, tess_lang, config=config)
                    cyr = sum(1 for l in lines for c in l.text if '\u0400' <= c <= '\u04FF')
                    if lines:
                        logger.info(f"Name-zone {name} PSM{psm}: {len(lines)} lines, {cyr} Cyrillic")

                    # Collect individual Cyrillic lines from all strategies
                    for l in lines:
                        line_cyr = sum(1 for c in l.text if '\u0400' <= c <= '\u04FF')
                        if line_cyr >= 2:
                            # Apply OCR corrections
                            corrected_text = self._correct_ocr_names(l.text.strip())
                            # Adjust coordinates back to full image
                            adjusted = OCRLine(
                                text=corrected_text,
                                confidence=l.confidence,
                                bbox=l.bbox,
                                center_x=l.center_x / scale + crop_x1,
                                center_y=l.center_y / scale + crop_y1
                            )
                            all_lines.append(adjusted)
                except Exception as e:
                    logger.warning(f"Name-zone {name} PSM{psm} failed: {e}")

        # ---- Also try per-word OCR with image_to_string PSM 8 (single word) ----
        # This can catch words that image_to_data misses entirely
        try:
            for name, processed in preprocessings[:5]:  # Use first 5 preprocessings
                pil_img = PILImage.fromarray(processed)
                raw_text = pytesseract.image_to_string(pil_img, lang=tess_lang, config="--psm 6")
                if raw_text:
                    for line_text in raw_text.strip().split("\n"):
                        clean = line_text.strip()
                        if not clean:
                            continue
                        cyr_count = sum(1 for c in clean if '\u0400' <= c <= '\u04FF')
                        if cyr_count >= 2 and len(clean) >= 2:
                            corrected = self._correct_ocr_names(clean)
                            logger.info(f"Name-zone string-mode {name}: '{clean}' → '{corrected}'")
                            # We don't know exact bbox, use approximate center
                            all_lines.append(OCRLine(
                                text=corrected,
                                confidence=0.5,
                                center_x=crop_x1 + crop_w / 2,
                                center_y=crop_y1 + crop_h / 2
                            ))
        except Exception as e:
            logger.warning(f"Name-zone string-mode failed: {e}")

        # Deduplicate: keep best version for each text/spatial region
        merged = []
        for line in all_lines:
            duplicate = False
            line_upper = line.text.strip().upper()
            for i, existing in enumerate(merged):
                existing_upper = existing.text.strip().upper()
                # Exact text match
                if line_upper == existing_upper:
                    duplicate = True
                    break
                # Same vertical position
                if abs(line.center_y - existing.center_y) < 25:
                    new_cyr = sum(1 for c in line.text if '\u0400' <= c <= '\u04FF')
                    old_cyr = sum(1 for c in existing.text if '\u0400' <= c <= '\u04FF')
                    if new_cyr > old_cyr:
                        merged[i] = line
                    duplicate = True
                    break
            if not duplicate:
                merged.append(line)

        merged.sort(key=lambda l: l.center_y)
        logger.info(f"Name-zone final: {len(merged)} unique lines")
        for l in merged:
            logger.info(f"  Name-zone: '{l.text}' conf={l.confidence:.0%} y={l.center_y:.0f}")
        return merged

    def _ocr_dl_name_zone(self, image, image_original=None) -> list:
        """Dedicated OCR pass for driving license name zone.
        DL names are in the upper-right portion of the card (fields 1-3).
        Crops that area, aggressively upscales, sharpens, and re-OCRs."""
        src = image_original if image_original is not None else image
        h, w = src.shape[:2]

        # DL names are roughly in the upper-right 65% x top 50%
        # (left side has the photo)
        crop_x1 = int(w * 0.30)
        crop_y1 = int(h * 0.10)
        crop_y2 = int(h * 0.55)
        name_crop = src[crop_y1:crop_y2, crop_x1:]
        crop_h, crop_w = name_crop.shape[:2]
        logger.info(f"DL name-zone crop: x={crop_x1}-{w}, y={crop_y1}-{crop_y2}, size={crop_w}x{crop_h}")

        # Aggressively upscale to at least 1500px wide
        target_w = 1500
        scale = max(1.0, target_w / max(crop_w, 1))
        scale = min(scale, 4.0)  # Don't go over 4x
        if scale > 1.0:
            name_crop = cv2.resize(name_crop, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
            logger.info(f"DL name-zone upscaled: {name_crop.shape[1]}x{name_crop.shape[0]} (scale={scale:.1f}x)")

        # Sharpen for better text edge detection
        kernel = np.array([[-1, -1, -1], [-1, 9, -1], [-1, -1, -1]])
        sharpened = cv2.filter2D(name_crop, -1, kernel)

        all_lines = []

        # Pass 1: EasyOCR on sharpened crop
        if HAS_EASYOCR and self.easyocr_reader:
            try:
                results = self.easyocr_reader.readtext(sharpened, detail=1, paragraph=False)
                for (bbox, txt, conf) in results:
                    if conf < 0.1:
                        continue
                    cx = (bbox[0][0] + bbox[2][0]) / 2 / scale + crop_x1
                    cy = (bbox[0][1] + bbox[2][1]) / 2 / scale + crop_y1
                    all_lines.append(OCRLine(text=txt.strip(), confidence=conf, center_x=cx, center_y=cy))
                    logger.info(f"DL name-zone EasyOCR: '{txt.strip()}' conf={conf:.2f}")
            except Exception as e:
                logger.warning(f"DL name-zone EasyOCR failed: {e}")

        # Pass 2: EasyOCR on grayscale CLAHE version
        if HAS_EASYOCR and self.easyocr_reader and HAS_CV2:
            try:
                gray = cv2.cvtColor(name_crop, cv2.COLOR_BGR2GRAY) if len(name_crop.shape) == 3 else name_crop
                clahe = cv2.createCLAHE(clipLimit=4.0, tileGridSize=(8, 8))
                enhanced = clahe.apply(gray)
                enhanced_bgr = cv2.cvtColor(enhanced, cv2.COLOR_GRAY2BGR)
                results = self.easyocr_reader.readtext(enhanced_bgr, detail=1, paragraph=False)
                for (bbox, txt, conf) in results:
                    if conf < 0.1:
                        continue
                    cx = (bbox[0][0] + bbox[2][0]) / 2 / scale + crop_x1
                    cy = (bbox[0][1] + bbox[2][1]) / 2 / scale + crop_y1
                    all_lines.append(OCRLine(text=txt.strip(), confidence=conf, center_x=cx, center_y=cy))
                    logger.info(f"DL name-zone CLAHE: '{txt.strip()}' conf={conf:.2f}")
            except Exception as e:
                logger.warning(f"DL name-zone CLAHE failed: {e}")

        # Pass 3: English-only EasyOCR for Latin text (D'IAKOVA, DINA ANDREEVNA)
        # The bilingual reader often garbles Latin text; English-only reads it cleanly
        if HAS_EASYOCR and hasattr(self, 'easyocr_reader_en') and self.easyocr_reader_en and HAS_CV2:
            try:
                results = self.easyocr_reader_en.readtext(sharpened, detail=1, paragraph=False)
                for (bbox, txt, conf) in results:
                    if conf < 0.15:
                        continue
                    cx = (bbox[0][0] + bbox[2][0]) / 2 / scale + crop_x1
                    cy = (bbox[0][1] + bbox[2][1]) / 2 / scale + crop_y1
                    all_lines.append(OCRLine(text=txt.strip(), confidence=conf, center_x=cx, center_y=cy))
                    logger.info(f"DL name-zone EasyOCR-EN: '{txt.strip()}' conf={conf:.2f}")
            except Exception as e:
                logger.warning(f"DL name-zone EasyOCR-EN failed: {e}")

        # Deduplicate: keep highest confidence for similar text
        merged = []
        for line in all_lines:
            line_upper = line.text.strip().upper()
            if len(line_upper) < 2:
                continue
            dup = False
            for i, ex in enumerate(merged):
                ex_upper = ex.text.strip().upper()
                if line_upper == ex_upper or (abs(line.center_y - ex.center_y) < 20 and abs(line.center_x - ex.center_x) < 50):
                    if line.confidence > ex.confidence:
                        merged[i] = line
                    dup = True
                    break
            if not dup:
                merged.append(line)

        merged.sort(key=lambda l: l.center_y)
        logger.info(f"DL name-zone final: {len(merged)} unique lines")
        for l in merged:
            logger.info(f"  DL name-zone: '{l.text}' conf={l.confidence:.0%} y={l.center_y:.0f}")
        return merged

    def _preprocess_document_image(self, image) -> tuple:
        """Comprehensive document image preprocessing for scanned documents.
        Returns: (preprocessed_image, quality_info_dict)

        Operations:
        - Auto-detect and correct document orientation
        - Perspective correction (dewarp)
        - Remove scanner artifacts (black borders, noise)
        - Auto-crop to document area
        - Normalize lighting (handle uneven illumination)
        - Assess image quality (blur, noise, resolution)
        """
        if not HAS_CV2:
            return image, {}

        quality_info = {
            "blur_score": 0.0,
            "noise_detected": False,
            "resolution_ok": True,
            "orientation_corrected": False,
            "perspective_corrected": False,
            "edge_artifacts_removed": False,
            "lighting_normalized": False,
        }

        h, w = image.shape[:2]

        # === 1. Auto-crop to remove black borders (scanner artifacts) ===
        try:
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if len(image.shape) == 3 else image
            _, binary = cv2.threshold(gray, 20, 255, cv2.THRESH_BINARY)

            # Crop left/right
            col_sums = np.sum(binary, axis=0)
            col_threshold = h * 0.1
            left, right = 0, w
            for i in range(w):
                if col_sums[i] > col_threshold:
                    left = i
                    break
            for i in range(w - 1, -1, -1):
                if col_sums[i] > col_threshold:
                    right = i + 1
                    break

            # Crop top/bottom
            row_sums = np.sum(binary, axis=1)
            row_threshold = w * 0.1
            top, bottom = 0, h
            for i in range(h):
                if row_sums[i] > row_threshold:
                    top = i
                    break
            for i in range(h - 1, -1, -1):
                if row_sums[i] > row_threshold:
                    bottom = i + 1
                    break

            # Apply crop if reasonable
            crop_margin = 0.2
            if (left < w * crop_margin and right > w * (1 - crop_margin) and
                top < h * crop_margin and bottom > h * (1 - crop_margin)):
                image = image[top:bottom, left:right]
                quality_info["edge_artifacts_removed"] = True
                logger.info(f"Cropped document to {image.shape[1]}x{image.shape[0]}")
        except Exception as e:
            logger.warning(f"Edge artifact removal failed: {e}")

        h, w = image.shape[:2]

        # === 2. Normalize lighting ===
        try:
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if len(image.shape) == 3 else image
            kernel_size = max(50, min(h, w) // 8)
            if kernel_size % 2 == 0:
                kernel_size += 1

            local_mean = cv2.blur(gray, (kernel_size, kernel_size))
            illumination_ratio = np.std(local_mean) / (np.mean(local_mean) + 1)

            if illumination_ratio > 0.1:
                kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (int(kernel_size * 0.6), int(kernel_size * 0.6)))
                background = cv2.morphologyEx(gray, cv2.MORPH_CLOSE, kernel)
                background = cv2.GaussianBlur(background, (kernel_size, kernel_size), 0)
                # Divide by background to remove uneven lighting, then rescale to full range
                normalized = cv2.divide(gray, background, scale=200)  # scale to ~200 target mean
                normalized = np.clip(normalized, 0, 255).astype(np.uint8)

                # Only apply if it improves the image (mean brightness in reasonable range)
                norm_mean = np.mean(normalized)
                if 80 < norm_mean < 220:
                    if len(image.shape) == 3:
                        image_lab = cv2.cvtColor(image, cv2.COLOR_BGR2LAB)
                        image_lab[:,:,0] = normalized
                        image = cv2.cvtColor(image_lab, cv2.COLOR_LAB2BGR)
                    else:
                        image = normalized
                    quality_info["lighting_normalized"] = True
                    logger.info(f"Lighting normalized (mean={norm_mean:.0f})")
                else:
                    logger.info(f"Lighting normalization skipped — result too dark/bright (mean={norm_mean:.0f})")
        except Exception as e:
            logger.warning(f"Lighting normalization failed: {e}")

        # === 3. Perspective correction ===
        try:
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if len(image.shape) == 3 else image
            edges = cv2.Canny(gray, 50, 150)
            kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (5, 5))
            edges = cv2.dilate(edges, kernel, iterations=2)

            contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            if len(contours) > 0:
                largest = max(contours, key=cv2.contourArea)
                area = cv2.contourArea(largest)

                if area > w * h * 0.3:
                    epsilon = 0.02 * cv2.arcLength(largest, True)
                    approx = cv2.approxPolyDP(largest, epsilon, True)

                    if len(approx) == 4:
                        src_pts = np.float32(approx.reshape(4, 2))
                        width = int(np.linalg.norm(src_pts[1] - src_pts[0]))
                        height = int(np.linalg.norm(src_pts[2] - src_pts[1]))

                        dst_pts = np.float32([
                            [0, 0],
                            [width, 0],
                            [width, height],
                            [0, height]
                        ])

                        matrix = cv2.getPerspectiveTransform(src_pts, dst_pts)
                        warped = cv2.warpPerspective(image, matrix, (width, height))

                        if warped.shape[0] > h * 0.5 and warped.shape[1] > w * 0.5:
                            image = warped
                            quality_info["perspective_corrected"] = True
                            logger.info(f"Perspective correction applied")
        except Exception as e:
            logger.warning(f"Perspective correction failed: {e}")

        # === 4. Quality assessment ===
        try:
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if len(image.shape) == 3 else image
            laplacian = cv2.Laplacian(gray, cv2.CV_64F)
            blur_score = laplacian.var()
            quality_info["blur_score"] = float(blur_score)

            if blur_score < 100:
                logger.warning(f"Image appears blurry (Laplacian variance={blur_score:.1f})")

            h, w = image.shape[:2]
            if w < 800 or h < 600:
                quality_info["resolution_ok"] = False
                logger.warning(f"Image resolution may be too low: {w}x{h}")

            if blur_score > 500:
                quality_info["noise_detected"] = True
        except Exception as e:
            logger.warning(f"Quality assessment failed: {e}")

        return image, quality_info

    def _assess_image_quality_detailed(self, image) -> dict:
        """Detailed image quality assessment for warnings/logging.
        Returns dict with: blur_score, noise_level, resolution, estimated_dpi"""
        quality = {
            "blur_score": 0.0,
            "noise_level": "unknown",
            "resolution": (0, 0),
            "estimated_dpi": 0,
        }

        if not HAS_CV2:
            return quality

        try:
            h, w = image.shape[:2]
            quality["resolution"] = (w, h)

            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if len(image.shape) == 3 else image

            # Blur score via Laplacian variance
            laplacian = cv2.Laplacian(gray, cv2.CV_64F)
            quality["blur_score"] = float(laplacian.var())

            # Noise classification based on blur_score
            bs = quality["blur_score"]
            if bs < 50:
                quality["noise_level"] = "very_blurry"
            elif bs < 100:
                quality["noise_level"] = "blurry"
            elif bs < 300:
                quality["noise_level"] = "normal"
            elif bs < 600:
                quality["noise_level"] = "sharp"
            else:
                quality["noise_level"] = "very_noisy"
        except Exception as e:
            logger.warning(f"Detailed quality assessment failed: {e}")

        return quality

    def _detect_and_correct_rotation(self, image) -> tuple:
        """Detect if document needs rotation. Test both original and rotated versions.
        Returns: (best_image, rotation_angle, rotated_flag)

        Note: Rotation detection is now simplified since PaddleOCR is no longer available.
        EasyOCR handles rotation reasonably well internally.
        """
        if not HAS_CV2:
            return image, 0, False

        # For now, rotation detection is disabled as PaddleOCR was the primary method.
        # EasyOCR handles most rotations internally.
        return image, 0, False

    def _pdf_render_page(self, doc, page_num: int, dpi: int = 300):
        """Render a single PDF page to OpenCV image at given DPI."""
        import fitz
        page = doc.load_page(page_num)
        mat = fitz.Matrix(dpi / 72, dpi / 72)
        pix = page.get_pixmap(matrix=mat)
        img_data = np.frombuffer(pix.samples, dtype=np.uint8)
        if pix.n == 4:
            img = img_data.reshape(pix.h, pix.w, 4)
            img = cv2.cvtColor(img, cv2.COLOR_RGBA2BGR)
        elif pix.n == 3:
            img = img_data.reshape(pix.h, pix.w, 3)
            img = cv2.cvtColor(img, cv2.COLOR_RGB2BGR)
        else:
            img = img_data.reshape(pix.h, pix.w)
            img = cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)
        return img

    def _pdf_prescreen_pages(self, pdf_bytes: bytes) -> list:
        """Two-phase pre-screen: Phase 1 uses fast image heuristics on ALL pages (no OCR).
        Phase 2 runs lightweight Tesseract OCR only on top 3 candidates.
        Returns list of (page_num, score) sorted by score descending."""
        from PIL import Image as PILImage
        try:
            import fitz
        except ImportError:
            return []

        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        num_pages = len(doc)
        logger.info(f"PDF pre-screening {num_pages} pages (Phase 1: image heuristics)...")

        # === PHASE 1: Fast image-only heuristics on all pages (NO OCR) ===
        phase1_scores = []
        for page_num in range(num_pages):
            img = self._pdf_render_page(doc, page_num, dpi=150)  # Lower DPI for speed
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            h, w = img.shape[:2]
            score = 0

            # 1. Content density
            _, binary = cv2.threshold(gray, 240, 255, cv2.THRESH_BINARY)
            non_white_ratio = cv2.countNonZero(binary) / (h * w)
            score += min(10, non_white_ratio * 100)

            # 2. Text line density
            sobelx = cv2.Sobel(gray, cv2.CV_64F, 1, 0, ksize=3)
            rows_with_text = np.sum(np.abs(sobelx) > 30, axis=1)
            text_line_density = np.sum(rows_with_text > w * 0.3) / h
            score += min(5, text_line_density * 20)

            # 3. Photo-like rectangular regions
            edges = cv2.Canny(gray, 50, 150)
            kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3))
            edges = cv2.dilate(edges, kernel, iterations=1)
            contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            rects = 0
            for c in contours:
                area = cv2.contourArea(c)
                if area < h * w * 0.05:
                    continue
                approx = cv2.approxPolyDP(c, 0.02 * cv2.arcLength(c, True), True)
                if len(approx) == 4:
                    bx, by, bw, bh = cv2.boundingRect(approx)
                    if 0.6 < bw / max(bh, 1) < 1.2:
                        rects += 1
            score += min(5, rects)

            # Penalize mostly-empty pages
            if non_white_ratio < 0.1:
                score -= 5

            phase1_scores.append((page_num, score))
            logger.info(f"PDF Phase1 page {page_num + 1}: score={score:.1f} (content={non_white_ratio:.0%}, text={text_line_density:.0%}, rects={rects})")
            del img, gray  # Free memory immediately

        # Sort and take top 3 candidates for Phase 2
        phase1_scores.sort(key=lambda x: -x[1])
        top_candidates = [p[0] for p in phase1_scores[:3]]
        logger.info(f"Phase 1 top candidates: pages {[p+1 for p in top_candidates]}")

        # === PHASE 2: Lightweight Tesseract OCR on top 3 candidates only ===
        # Uses Tesseract (fast, low memory) instead of EasyOCR to save RAM for actual processing
        page_scores = []
        for page_num in top_candidates:
            img = self._pdf_render_page(doc, page_num, dpi=200)
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8, 8))
            enhanced = clahe.apply(gray)

            text = ""
            if HAS_TESSERACT:
                try:
                    pil_img = PILImage.fromarray(enhanced)
                    text = pytesseract.image_to_string(pil_img, lang=self._detect_tess_lang(), config="--psm 6")
                except Exception:
                    text = ""

            # Score based on document indicators
            ocr_score = 0
            dates = re.findall(r'\d{2}[.\-]\d{2}[.\-]\d{4}', text)
            ocr_score += len(dates) * 3
            if re.search(r'\d{2}\s+\d{2}\s+\d{6}', text):
                ocr_score += 5
            if re.search(r'\d{3}\s*[-–.]\s*\d{3}', text):
                ocr_score += 3
            if re.search(r'(?i)(УФМС|Y.?MC|МВД|MVD|УМВД)', text):
                ocr_score += 3
            if re.search(r'(?i)(фамил|ФАМИЛ|имя|ИМЯ|отчеств|ОТЧЕСТВ)', text):
                ocr_score += 4
            if re.search(r'(?i)(МУЖ|ЖЕН|MYX|MУЖ)', text):
                ocr_score += 3
            if re.search(r'(?i)(РОССИЙ|POCCИЙ|ФЕДЕРАЦ)', text):
                ocr_score += 2
            if re.search(r'(?i)(ОБЛАСТЬ|РАЙОН)', text):
                ocr_score += 2
            if re.search(r'(?i)(ТЕРРИТОР|ПУНКТ|ПОДРАЗД)', text):
                ocr_score += 2
            if re.search(r'(?i)(ПАСПОРТ|NACNOPT)', text):
                ocr_score += 2
            if re.search(r'(?i)(ВОДИТЕЛЬСК|DRIVING|ГИБДД)', text):
                ocr_score += 4
            if re.search(r'(?i)(СНИЛС|СТРАХОВОЕ|ИНН\s*:?\s*\d)', text):
                ocr_score += 4
            cyr_count = sum(1 for c in text if '\u0400' <= c <= '\u04FF')
            ocr_score += min(5, cyr_count // 50)
            if len(text.strip()) < 50:
                ocr_score -= 3

            # Get phase1 image score for this page
            img_score = next(s for p, s in phase1_scores if p == page_num)
            total_score = img_score * 0.4 + ocr_score * 0.6

            page_scores.append((page_num, total_score))
            logger.info(f"PDF Phase2 page {page_num + 1}: total={total_score:.1f} (image={img_score:.1f}, ocr={ocr_score:.1f}, dates={len(dates)}, cyr={cyr_count})")
            del img, gray, enhanced

        doc.close()
        page_scores.sort(key=lambda x: -x[1])
        return page_scores

    def _pdf_to_images(self, pdf_bytes: bytes, page_nums: list = None) -> list:
        """Convert specific PDF pages to OpenCV images.
        If page_nums is None, converts all pages.
        Returns list of (page_num, cv2_image)."""
        images = []
        # Try PyMuPDF (fitz) first
        try:
            import fitz
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            num_pages = len(doc)
            target_pages = page_nums if page_nums is not None else list(range(num_pages))
            logger.info(f"PDF rendering {len(target_pages)} of {num_pages} pages at 300 DPI")
            for page_num in target_pages:
                if 0 <= page_num < num_pages:
                    img = self._pdf_render_page(doc, page_num, dpi=300)
                    images.append((page_num, img))
                    logger.info(f"PDF page {page_num + 1}: {img.shape[1]}x{img.shape[0]}")
            doc.close()
            return images
        except ImportError:
            logger.info("PyMuPDF not available, trying pdf2image")
        except Exception as e:
            logger.warning(f"PyMuPDF failed: {e}")

        # Fallback: pdf2image (requires poppler)
        try:
            from pdf2image import convert_from_bytes
            pil_images = convert_from_bytes(pdf_bytes, dpi=300)
            logger.info(f"PDF opened with pdf2image: {len(pil_images)} pages")
            target_pages = page_nums if page_nums is not None else list(range(len(pil_images)))
            for page_num in target_pages:
                if 0 <= page_num < len(pil_images):
                    img = cv2.cvtColor(np.array(pil_images[page_num]), cv2.COLOR_RGB2BGR)
                    images.append((page_num, img))
            return images
        except ImportError:
            logger.warning("pdf2image not available either")
        except Exception as e:
            logger.warning(f"pdf2image failed: {e}")

        return images

    def _is_pdf(self, data: bytes) -> bool:
        """Check if data is a PDF file."""
        return data[:5] == b'%PDF-'

    def _detect_bank_statement_format(self, data: bytes) -> tuple[Optional[str], Optional[dict]]:
        """Detect bank statement format (txt_1c, csv, or xlsx).
        Returns (format_type, parsed_data) or (None, None) if not detected."""

        # Check for XLSX first (PK magic bytes)
        if data[:2] == b'PK':
            try:
                from io import BytesIO
                xlsx_file = BytesIO(data)
                if HAS_OPENPYXL:
                    wb = openpyxl.load_workbook(xlsx_file)
                    if 'Выписка' in wb.sheetnames:
                        logger.info("Detected XLSX bank statement with 'Выписка' sheet")
                        return ('xlsx', {'data': data})
            except Exception as e:
                logger.debug(f"XLSX detection failed: {e}")

        # Try to decode as text and check for TXT/CSV formats
        # Try Windows-1251 FIRST (most common for Russian bank files), then UTF-8
        for encoding in ['windows-1251', 'utf-8']:
            try:
                text = data.decode(encoding)  # strict decode, no errors='ignore'
                lines = text.split('\n', 10)

                # Check for 1C TXT format
                first_line = lines[0].strip() if lines else ''
                if '1CClientBankExchange' in first_line:
                    # Verify encoding is correct: check if Cyrillic keys are readable
                    has_cyrillic = any('Версия' in l or 'Кодировка' in l or 'Дата' in l for l in lines[:10])
                    if has_cyrillic or encoding == 'windows-1251':
                        logger.info(f"Detected 1C TXT bank statement (encoding: {encoding})")
                        return ('txt_1c', {'data': data, 'encoding': encoding})

                # Check for CSV format
                if 'Счет;Дата операции' in first_line or 'Код операции' in first_line:
                    logger.info(f"Detected CSV bank statement (encoding: {encoding})")
                    return ('csv', {'data': data, 'encoding': encoding})
            except (UnicodeDecodeError, Exception) as e:
                logger.debug(f"Text decode with {encoding} failed: {e}")
                continue

        return None, None

    def process(self, image_bytes: bytes, document_type_hint: str = None) -> dict:
        start = time.time()
        result = {
            "status": "completed",
            "document_type": "unknown",
            "classification_confidence": 0.0,
            "overall_confidence": 0.0,
            "fields": {},
            "validation": {},
            "warnings": [],
            "processing_time_ms": 0,
        }

        try:
            # Step 0: Check if input is PDF — extract page images first
            if self._is_pdf(image_bytes):
                logger.info("Input is PDF — extracting pages as images")

                # Pre-screen pages to find the ones most likely containing document data
                page_scores = self._pdf_prescreen_pages(image_bytes)
                if not page_scores:
                    # Pre-screening failed — try all pages (but limit to 5)
                    pdf_pages = self._pdf_to_images(image_bytes)
                    if pdf_pages and len(pdf_pages) > 5:
                        pdf_pages = pdf_pages[:5]
                    target_page_nums = [p[0] for p in (pdf_pages or [])]
                else:
                    # Take top 3 pages by pre-screen score (or fewer if PDF is small)
                    top_pages = [p for p in page_scores if p[1] > 0][:3]
                    if not top_pages:
                        # No page scored above 0 — take first 3
                        top_pages = page_scores[:3]
                    target_page_nums = [p[0] for p in top_pages]
                    logger.info(f"PDF pre-screen: processing pages {[p+1 for p in target_page_nums]} (scores: {[(p[0]+1, p[1]) for p in page_scores[:5]]})")

                # Render only the selected pages at full 300 DPI
                pdf_pages = self._pdf_to_images(image_bytes, page_nums=target_page_nums)
                if not pdf_pages:
                    result["status"] = "failed"
                    result["warnings"].append("Не удалось извлечь страницы из PDF. Установите PyMuPDF: pip install PyMuPDF")
                    result["processing_time_ms"] = int((time.time() - start) * 1000)
                    return result

                # Process selected pages and pick the best result
                best_result = None
                best_score = -1
                total_pages_in_pdf = len(page_scores) if page_scores else len(pdf_pages)
                for page_num, page_img in pdf_pages:
                    logger.info(f"Full processing PDF page {page_num + 1}...")
                    # Convert page image to bytes for recursive processing
                    _, page_encoded = cv2.imencode(".png", page_img)
                    page_bytes = page_encoded.tobytes()
                    page_result = self.process(page_bytes, document_type_hint=document_type_hint)
                    page_conf = page_result.get("overall_confidence", 0)
                    num_fields = len(page_result.get("fields", {}))
                    # Score: heavily weight number of fields + name presence
                    has_name = 1 if "last_name" in page_result.get("fields", {}) else 0
                    has_number = 1 if "number" in page_result.get("fields", {}) else 0
                    has_date = 1 if "birth_date" in page_result.get("fields", {}) else 0
                    score = num_fields * 0.15 + page_conf * 0.3 + has_name * 0.3 + has_number * 0.15 + has_date * 0.1
                    logger.info(f"PDF page {page_num + 1}: type={page_result.get('document_type')}, "
                               f"conf={page_conf:.2f}, fields={num_fields}, name={has_name}, score={score:.2f}")
                    if score > best_score:
                        best_score = score
                        best_result = page_result
                        best_result["_pdf_page"] = page_num

                if best_result:
                    best_result["processing_time_ms"] = int((time.time() - start) * 1000)
                    best_page = best_result.get('_pdf_page', 0) + 1
                    best_result["warnings"].insert(0, f"PDF: {total_pages_in_pdf} стр., обработано {len(pdf_pages)}, лучший результат со стр. {best_page}")
                    best_result.pop("_pdf_page", None)
                    return best_result
                else:
                    result["status"] = "failed"
                    result["warnings"].append("PDF: не удалось распознать ни одну страницу")
                    result["processing_time_ms"] = int((time.time() - start) * 1000)
                    return result

            # Step 0.5: Check for bank statement (structured data file — bypass OCR)
            bank_fmt, bank_data = self._detect_bank_statement_format(image_bytes)
            if bank_fmt:
                logger.info(f"Detected bank statement format: {bank_fmt}")
                try:
                    fields = self._parse_bank_statement(image_bytes, bank_fmt, bank_data)
                    result["document_type"] = "bank_statement"
                    result["classification_confidence"] = 1.0
                    result["overall_confidence"] = 1.0
                    result["fields"] = {}
                    for name, f in fields.items():
                        if name == 'transactions' and isinstance(f.value, list):
                            result["fields"][name] = {
                                "value": f.value,
                                "confidence": 1.0,
                                "source": "bank_statement_parsing",
                                "auto_fill": True,
                            }
                        else:
                            result["fields"][name] = {
                                "value": str(f.value),
                                "confidence": float(round(f.confidence, 3)),
                                "source": str(f.source),
                                "auto_fill": bool(f.confidence >= 0.95),
                            }
                    result["processing_time_ms"] = int((time.time() - start) * 1000)
                    return result
                except Exception as e:
                    logger.warning(f"Bank statement parsing failed: {e}")
                    import traceback
                    logger.warning(traceback.format_exc())
                    result["warnings"].append(f"Ошибка при обработке банковской выписки: {e}")
                    result["status"] = "failed"
                    result["processing_time_ms"] = int((time.time() - start) * 1000)
                    return result  # Don't fall through to image decode for structured files

            # Step 1: Decode image (with HEIC/HEIF support)
            if HAS_CV2:
                nparr = np.frombuffer(image_bytes, np.uint8)
                image = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

                # If cv2 can't decode (e.g. HEIC/HEIF), try pillow-heif directly then Pillow
                if image is None:
                    try:
                        from io import BytesIO as BIO
                        decoded = False
                        # Try pillow-heif directly (without register_heif_opener to avoid mode setter bug)
                        try:
                            import pillow_heif
                            heif_image = pillow_heif.open_heif(BIO(image_bytes))
                            arr = np.asarray(heif_image)
                            if len(arr.shape) == 3 and arr.shape[2] == 4:
                                image = cv2.cvtColor(arr, cv2.COLOR_RGBA2BGR)
                            elif len(arr.shape) == 3 and arr.shape[2] == 3:
                                image = cv2.cvtColor(arr, cv2.COLOR_RGB2BGR)
                            else:
                                image = arr
                            decoded = True
                            logger.info(f"Decoded via pillow-heif: {image.shape[1]}x{image.shape[0]}")
                        except Exception:
                            pass
                        # Fallback to standard Pillow (for BMP, TIFF, etc.)
                        if not decoded:
                            from PIL import Image as PILImage
                            pimg = PILImage.open(BIO(image_bytes))
                            pimg = pimg.convert("RGB")
                            image = cv2.cvtColor(np.array(pimg), cv2.COLOR_RGB2BGR)
                            logger.info(f"Decoded via Pillow: {image.shape[1]}x{image.shape[0]}")
                    except Exception as e2:
                        logger.error(f"Image decode failed: {e2}")
                        result["warnings"].append(f"Cannot decode image: {e2}")
                        result["status"] = "failed"
                        return result

                h, w = image.shape[:2]
                logger.info(f"Image loaded: {w}x{h}")

                # Save ORIGINAL image for EasyOCR (it has its own internal preprocessing)
                # EasyOCR works best with clean, minimally-processed images
                # For small images (DL cards, small photos), UPSCALE for better OCR
                # For large images, downscale to save memory
                min_dim_for_ocr = 1400  # Minimum dimension for good OCR quality
                max_orig = 2500  # Keep more resolution for EasyOCR
                if max(h, w) < min_dim_for_ocr:
                    # Small image (likely a card document) — upscale for better recognition
                    scale_up = min_dim_for_ocr / max(h, w)
                    # Don't upscale more than 3x
                    scale_up = min(scale_up, 3.0)
                    image_original = cv2.resize(image, (int(w * scale_up), int(h * scale_up)), interpolation=cv2.INTER_CUBIC)
                    logger.info(f"Upscaled small image for EasyOCR: {w}x{h} → {image_original.shape[1]}x{image_original.shape[0]} (scale={scale_up:.1f}x)")
                elif max(h, w) > max_orig:
                    scale_orig = max_orig / max(h, w)
                    image_original = cv2.resize(image, (int(w * scale_orig), int(h * scale_orig)))
                    logger.info(f"Downscaled image for EasyOCR: {image_original.shape[1]}x{image_original.shape[0]}")
                else:
                    image_original = image.copy()

                # === Preprocessing for Tesseract (heavier processing needed) ===
                image, preproc_info = self._preprocess_document_image(image)
                h, w = image.shape[:2]
                logger.info(f"After preprocessing: {w}x{h}, "
                           f"edge_artifacts_removed={preproc_info.get('edge_artifacts_removed', False)}, "
                           f"perspective_corrected={preproc_info.get('perspective_corrected', False)}, "
                           f"lighting_normalized={preproc_info.get('lighting_normalized', False)}")

                # Size limiting
                if max(h, w) > 3000:
                    scale = 3000 / max(h, w)
                    image = cv2.resize(image, (int(w * scale), int(h * scale)))
                    h, w = image.shape[:2]
                    logger.info(f"Resized to {w}x{h}")

                # Auto-brightness correction for dark photos (Tesseract version)
                gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
                mean_brightness = np.mean(gray)
                logger.info(f"Image mean brightness: {mean_brightness:.0f}/255")
                if mean_brightness < 140:
                    alpha = min(2.5, 170.0 / max(mean_brightness, 1))
                    beta = max(0, 60 - mean_brightness * 0.3)
                    image = cv2.convertScaleAbs(image, alpha=alpha, beta=beta)
                    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
                    logger.info(f"Brightness corrected: alpha={alpha:.2f}, beta={beta:.1f}, new mean={np.mean(gray):.0f}")

                # Enhance for Tesseract OCR via CLAHE
                clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8, 8))
                enhanced_gray = clahe.apply(gray)
                image_enhanced = cv2.cvtColor(enhanced_gray, cv2.COLOR_GRAY2BGR)
            else:
                image = image_bytes
                image_enhanced = image_bytes
                h, w = 0, 0

            # Step 2: OCR
            ocr_lines = []
            if HAS_CV2:
                # Priority: Use ensemble (EasyOCR + Tesseract) if EasyOCR is available
                # EasyOCR is PRIMARY for Russian text, especially soft signs (ДЬ)
                if self.easyocr_reader:
                    logger.info("Using ensemble OCR (EasyOCR PRIMARY + Tesseract SECONDARY)")
                    ocr_lines = self._ocr_ensemble(image, image_original=image_original)
                elif self.engine == "tesseract":
                    # Check if Russian language is available
                    tess_lang = self._detect_tess_lang()
                    if "rus" not in tess_lang:
                        result["warnings"].append("КРИТИЧНО: Русский язык не установлен в Tesseract! ФИО и другие текстовые поля не будут распознаны. Выполните команду: brew install tesseract-lang")
                    if self.tessdata_quality == "fast":
                        result["warnings"].append("Используется быстрая модель Tesseract (tessdata_fast) — качество распознавания ФИО будет низким. Рекомендуется установить tessdata_best для русского языка.")

                    # Fallback to Tesseract only
                    logger.info("Using Tesseract multi-strategy OCR only")
                    ocr_lines = self._ocr_tesseract_multi_strategy(image)
                    if not ocr_lines:
                        ocr_lines = self._ocr_tesseract(image_enhanced)
                    logger.info(f"OCR extracted {len(ocr_lines)} lines")
                else:
                    result["warnings"].append("No OCR engine available — returning demo result")
                    return self._demo_result(result, start)
            else:
                result["warnings"].append("No OCR engine available — returning demo result")
                return self._demo_result(result, start)

            ocr_lines.sort(key=lambda l: (round(l.center_y / 20) * 20, l.center_x))

            if not ocr_lines:
                result["warnings"].append("No text detected")
                result["status"] = "failed"
                return result

            full_text = "\n".join(l.text for l in ocr_lines)
            logger.info(f"OCR full text ({len(ocr_lines)} lines):\n{full_text[:500]}")

            # === NEW: Image quality assessment ===
            if HAS_CV2:
                try:
                    quality_info = self._assess_image_quality_detailed(image)
                    blur_score = quality_info.get("blur_score", 0.0)
                    noise_level = quality_info.get("noise_level", "unknown")
                    resolution = quality_info.get("resolution", (0, 0))

                    if blur_score < 50:
                        result["warnings"].append(f"Внимание: изображение сильно размыто (Laplacian={blur_score:.1f})")
                    elif blur_score < 100:
                        result["warnings"].append(f"Внимание: изображение размыто (Laplacian={blur_score:.1f})")

                    logger.info(f"Image quality: blur={blur_score:.1f}, noise={noise_level}, resolution={resolution[0]}x{resolution[1]}")
                except Exception as e:
                    logger.warning(f"Quality assessment failed: {e}")

            # Step 2b: Dedicated MRZ OCR pass (bottom of image, binarized)
            dedicated_mrz_lines = []
            if HAS_CV2:
                try:
                    h, w = image.shape[:2]
                    mrz_crop = image[int(h * 0.60):, :]
                    # Preprocess MRZ zone: convert to grayscale
                    if len(mrz_crop.shape) == 3:
                        mrz_gray = cv2.cvtColor(mrz_crop, cv2.COLOR_BGR2GRAY)
                    else:
                        mrz_gray = mrz_crop

                    # Upscale MRZ zone for better OCR accuracy
                    mrz_scale = max(1.0, 1500 / max(mrz_gray.shape[1], 1))
                    if mrz_scale > 1.0:
                        mrz_gray = cv2.resize(mrz_gray, None, fx=mrz_scale, fy=mrz_scale, interpolation=cv2.INTER_CUBIC)
                        logger.info(f"MRZ zone upscaled: {mrz_gray.shape[1]}x{mrz_gray.shape[0]} (scale={mrz_scale:.2f}x)")

                    # FIRST TRY: English-only EasyOCR reader (best for MRZ, avoids Cyrillic misrecognition)
                    if HAS_EASYOCR and self.easyocr_reader_en and not dedicated_mrz_lines:
                        try:
                            logger.info("Attempting MRZ OCR with English-only EasyOCR reader...")
                            # Try adaptive threshold first
                            mrz_bin = cv2.adaptiveThreshold(mrz_gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 31, 10)
                            mrz_bin_bgr = cv2.cvtColor(mrz_bin, cv2.COLOR_GRAY2BGR)
                            mrz_results = self.easyocr_reader_en.readtext(mrz_bin_bgr, detail=1, paragraph=False)
                            for (bbox, txt, conf) in mrz_results:
                                clean = txt.replace(" ", "").upper()
                                # No need for Cyrillic mapping with English-only reader
                                # Replace common OCR noise as < filler
                                clean = re.sub(r"[«»{}\[\]()~_\-=.,;:!?|/\\]", "<", clean)
                                clean = self._normalize_mrz_k_to_filler(clean)
                                logger.info(f"EasyOCR-EN MRZ raw: '{txt}' → cleaned: '{clean[:50]}' (conf={conf:.0%})")
                                if len(clean) >= 28 and re.match(r"^[A-Z0-9<]+$", clean):
                                    dedicated_mrz_lines.append(clean)
                                    logger.info(f"EasyOCR-EN MRZ accepted: {clean[:50]}... (conf={conf:.0%})")

                            # If adaptive threshold didn't yield results, try Otsu binarization
                            if not dedicated_mrz_lines:
                                logger.info("Adaptive threshold produced no results, trying Otsu binarization...")
                                _, mrz_bin_otsu = cv2.threshold(mrz_gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                                # Also try inverted
                                for otsu_bin in [mrz_bin_otsu, cv2.bitwise_not(mrz_bin_otsu)]:
                                    mrz_bin_bgr = cv2.cvtColor(otsu_bin, cv2.COLOR_GRAY2BGR)
                                    mrz_results = self.easyocr_reader_en.readtext(mrz_bin_bgr, detail=1, paragraph=False)
                                    for (bbox, txt, conf) in mrz_results:
                                        clean = txt.replace(" ", "").upper()
                                        clean = re.sub(r"[«»{}\[\]()~_\-=.,;:!?|/\\]", "<", clean)
                                        clean = self._normalize_mrz_k_to_filler(clean)
                                        logger.info(f"EasyOCR-EN Otsu MRZ raw: '{txt}' → cleaned: '{clean[:50]}' (conf={conf:.0%})")
                                        if len(clean) >= 28 and re.match(r"^[A-Z0-9<]+$", clean):
                                            dedicated_mrz_lines.append(clean)
                                            logger.info(f"EasyOCR-EN Otsu MRZ accepted: {clean[:50]}... (conf={conf:.0%})")
                                    if len(dedicated_mrz_lines) >= 2:
                                        break
                        except Exception as e_en:
                            logger.warning(f"English-only EasyOCR MRZ pass failed: {e_en}")

                    # SECOND TRY: Tesseract MRZ (works well, eng-only) — try even if primary is EasyOCR
                    if HAS_TESSERACT and not dedicated_mrz_lines:
                        try:
                            logger.info("Attempting MRZ OCR with Tesseract...")
                            dedicated_mrz_lines = self._ocr_mrz_zone(image)
                        except Exception as e_tess:
                            logger.warning(f"Tesseract MRZ pass failed: {e_tess}")

                    # THIRD TRY: Fallback to bilingual EasyOCR reader (last resort)
                    if HAS_EASYOCR and self.easyocr_reader and not dedicated_mrz_lines:
                        try:
                            logger.info("Attempting MRZ OCR with bilingual EasyOCR reader (fallback)...")
                            # Try adaptive threshold
                            mrz_bin = cv2.adaptiveThreshold(mrz_gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 31, 10)
                            mrz_bin_bgr = cv2.cvtColor(mrz_bin, cv2.COLOR_GRAY2BGR)
                            mrz_results = self.easyocr_reader.readtext(mrz_bin_bgr, detail=1, paragraph=False)
                            for (bbox, txt, conf) in mrz_results:
                                clean = txt.replace(" ", "").upper()
                                # Map Cyrillic lookalikes to Latin (EasyOCR reads MRZ as Cyrillic)
                                cyr_to_lat = str.maketrans("АВСЕНКМОРТХУЗ", "ABCEHKMOPTXUZ")
                                clean = clean.translate(cyr_to_lat)
                                # Replace common OCR noise as < filler
                                clean = re.sub(r"[«»{}\[\]()~_\-=.,;:!?|/\\КкСсcC]", "<", clean)
                                # Remove remaining Cyrillic
                                clean = re.sub(r"[а-яёА-ЯЁ]", "", clean)
                                clean = self._normalize_mrz_k_to_filler(clean)
                                logger.info(f"EasyOCR bilingual MRZ raw: '{txt}' → cleaned: '{clean[:50]}' (conf={conf:.0%})")
                                if len(clean) >= 28 and re.match(r"^[A-Z0-9<]+$", clean):
                                    dedicated_mrz_lines.append(clean)
                                    logger.info(f"EasyOCR bilingual MRZ accepted: {clean[:50]}... (conf={conf:.0%})")
                        except Exception as e_bi:
                            logger.warning(f"Bilingual EasyOCR MRZ fallback failed: {e_bi}")

                    logger.info(f"Dedicated MRZ pass found {len(dedicated_mrz_lines)} lines")
                except Exception as e:
                    logger.warning(f"Dedicated MRZ pass failed: {e}")

            # Step 3: Classify
            doc_type, cls_conf = self._classify(full_text, document_type_hint)
            result["document_type"] = str(doc_type)
            result["classification_confidence"] = float(round(cls_conf, 3))

            # Step 3b: Dedicated name-zone OCR for internal passports
            # Crop the right-center area where names are and OCR with aggressive settings
            dedicated_name_lines = []
            if doc_type == "passport_rf" and self.engine == "tesseract" and HAS_CV2 and HAS_TESSERACT:
                try:
                    dedicated_name_lines = self._ocr_name_zone(image)
                    if dedicated_name_lines:
                        logger.info(f"Dedicated name-zone OCR found {len(dedicated_name_lines)} lines")
                        # Add unique name-zone lines to main OCR lines
                        existing_texts = {l.text.strip().upper() for l in ocr_lines}
                        for nl in dedicated_name_lines:
                            if nl.text.strip().upper() not in existing_texts:
                                ocr_lines.append(nl)
                                existing_texts.add(nl.text.strip().upper())
                                logger.info(f"Name-zone added: '{nl.text}' conf={nl.confidence:.0%} y={nl.center_y:.0f}")
                        ocr_lines.sort(key=lambda l: (round(l.center_y / 20) * 20, l.center_x))
                        full_text = "\n".join(l.text for l in ocr_lines)
                except Exception as e:
                    logger.warning(f"Dedicated name-zone OCR failed: {e}")

            # Step 3c: Dedicated name-zone OCR for driving licenses
            # DL has very small text — crop name area and re-OCR with upscale + sharpening
            if doc_type == "driver_license" and HAS_CV2 and HAS_EASYOCR and self.easyocr_reader:
                try:
                    dl_name_lines = self._ocr_dl_name_zone(image, image_original=image_original)
                    if dl_name_lines:
                        logger.info(f"DL name-zone OCR found {len(dl_name_lines)} lines")
                        existing_texts = {l.text.strip().upper() for l in ocr_lines}
                        for nl in dl_name_lines:
                            if nl.text.strip().upper() not in existing_texts:
                                ocr_lines.append(nl)
                                existing_texts.add(nl.text.strip().upper())
                                logger.info(f"DL name-zone added: '{nl.text}' conf={nl.confidence:.0%} y={nl.center_y:.0f}")
                        ocr_lines.sort(key=lambda l: (round(l.center_y / 20) * 20, l.center_x))
                        full_text = "\n".join(l.text for l in ocr_lines)
                except Exception as e:
                    logger.warning(f"DL name-zone OCR failed: {e}")

            # Step 4: Parse fields
            fields = self._parse_fields(doc_type, ocr_lines, full_text, dedicated_mrz_lines)
            result["fields"] = {
                name: {
                    "value": str(f.value),
                    "confidence": float(round(f.confidence, 3)),
                    "source": str(f.source),
                    "auto_fill": bool(f.confidence >= 0.95),
                    "needs_review": bool(0.70 <= f.confidence < 0.95),
                }
                for name, f in fields.items()
            }

            # Step 5: Validate
            result["validation"] = sanitize_for_json(self._validate(doc_type, result["fields"]))

            # Warnings from validation
            for fname, vres in result["validation"].items():
                if not vres.get("valid", True):
                    result["warnings"].append(f"{fname}: {vres.get('message', 'invalid')}")

            # Overall confidence
            if result["fields"]:
                confs = [float(f["confidence"]) for f in result["fields"].values()]
                result["overall_confidence"] = float(round(sum(confs) / len(confs), 3))

        except Exception as e:
            import traceback
            tb = traceback.format_exc()
            logger.exception(f"Pipeline error: {e}\n{tb}")
            result["status"] = "failed"
            result["warnings"].append(f"Ошибка обработки: {str(e)}")
            result["warnings"].append(f"Traceback: {tb[-500:]}")

        result["processing_time_ms"] = int((time.time() - start) * 1000)
        # Final sanitization: ensure all types are JSON-safe
        result = sanitize_for_json(result)
        return result

    def _classify(self, text: str, hint: str = None) -> tuple[str, float]:
        """Rule-based document classification."""
        scores = {}

        # Detect MRZ country code
        mrz_country = ""
        for mline in text.split("\n"):
            clean = mline.replace(" ", "")
            if len(clean) >= 30 and re.match(r"^[A-Z0-9<]+$", clean) and clean.startswith("P"):
                mrz_country = clean[2:5].replace("<", "")
                break

        # Passport RF (internal + загранпаспорт)
        score = 0
        # Cyrillic + common OCR misreads of "ПАСПОРТ" / "РОССИЙСКАЯ"
        if re.search(r"(?i)паспорт|NACNOPT|PASPORT|PASSPORT|HACNOPT|ПАСП0РТ", text): score += 3
        if re.search(r"(?i)российская|POCCUNCK|RUSSIAN|ФЕДЕРАЦ|FEDERAC|poCCH|РОССИ", text): score += 3
        if re.search(r"(?i)фамилия|surname|фамил", text): score += 2
        if re.search(r"(?i)отдел.*(УФМС|УМВД|МВД|полиции)", text): score += 3
        if re.search(r"\d{2}\s+\d{2}\s+\d{6}", text): score += 3  # internal passport series+number
        if re.search(r"\d{2}\s?\d{2}\s?\d{6}", text): score += 2
        if re.search(r"(?i)отчеств|patronym|given.?name", text): score += 2
        if re.search(r"\d{3}\s*[-–.]\s*\d{3}", text): score += 2
        if mrz_country == "RUS": score += 6
        if re.search(r"P\s*<\s*RUS", text): score += 5
        if re.search(r"(?i)\bМВД\b|\bMVD\b", text): score += 2
        scores["passport_rf"] = score / 27

        # Passport CIS (NOT Russia)
        score = 0
        if re.search(r"(?i)passeport|passport", text): score += 2
        if re.search(r"(?i)(O.?ZBEKISTON|QIRG.?IZSTAN|ТОҶИКИСТОН|ҚАЗАҚСТАН|UKRAINE|MOLDOVA)", text): score += 5
        if mrz_country and mrz_country != "RUS": score += 5
        if mrz_country == "RUS": score -= 10
        scores["passport_cis"] = max(0, score) / 12

        # Driver License
        score = 0
        if re.search(r"(?i)водительское\s*удостоверени", text): score += 5
        if re.search(r"(?i)водительск", text): score += 3
        if re.search(r"(?i)DRIVING\s*LICEN[CS]E", text): score += 4
        if re.search(r"(?i)PERMIS\s*DE\s*CONDUIRE", text): score += 4
        if re.search(r"(?i)ГИБДД|ГАИ|GIBDD", text): score += 4
        if re.search(r"(?i)категор|categor", text): score += 2
        if re.search(r"\b\d{2}\s+\d{2}\s+\d{6}\b", text): score += 3  # DL number format
        if re.search(r"(?i)4[аaа]\s*\)|4\s*[аa]\)", text): score += 2  # "4a)" label
        if re.search(r"(?i)place\s*of\s*birth|lieu\s*de\s*naissance", text): score += 2
        if re.search(r"(?i)4[bc]\)|4\s*[bc]\)", text): score += 2  # "4b)" or "4c)" labels
        if re.search(r"(?i)date\s*of\s*issue|date\s*of\s*expiry", text): score += 2
        scores["driver_license"] = score / 25

        # SNILS
        score = 0
        if re.search(r"\d{3}[-\s]?\d{3}[-\s]?\d{3}[-\s]?\d{2}", text): score += 5
        if re.search(r"(?i)страховое\s+свидетельство", text): score += 4
        if re.search(r"(?i)СНИЛС", text): score += 5
        scores["snils"] = score / 14

        # INN
        score = 0
        if re.search(r"(?i)свидетельство.*постановке", text): score += 4
        if re.search(r"(?i)ИНН\s*:?\s*\d{10,12}", text): score += 5
        if re.search(r"(?i)налоговом\s+органе", text): score += 4
        scores["inn"] = score / 13

        # Bank Statement
        score = 0
        if re.search(r"1CClientBankExchange", text): score += 10
        if re.search(r"ВерсияФормата", text): score += 3
        if re.search(r"СекцияДокумент|Платежное\s+поручение|Банковский\s+ордер", text): score += 5
        if re.search(r"НачальныйОстаток|КонечныйОстаток", text): score += 3
        if re.search(r"РасчСчет", text): score += 3
        if re.search(r"выписка\s*по\s*счету", text, re.IGNORECASE): score += 3
        if re.search(r"ПлательщикСчет|ПолучательСчет", text): score += 3
        if re.search(r"ПлательщикИНН|ПолучательИНН", text): score += 2
        if re.search(r"ПлательщикБИК", text): score += 2
        scores["bank_statement"] = score / 34

        # Map aliases to canonical types
        hint_aliases = {
            "passport_int": "passport_rf",
            "zagran": "passport_rf",
            "foreign_passport": "passport_rf",
        }
        if hint and hint in hint_aliases:
            hint = hint_aliases[hint]

        if hint and hint in scores:
            scores[hint] = min(1.0, scores[hint] + 0.4)

        best = max(scores, key=scores.get)
        # If user provided a hint, always trust it (minimum 0.5 confidence)
        if hint and hint in scores:
            return (hint, max(scores[hint], 0.5))
        return (best, scores[best]) if scores[best] >= 0.15 else ("unknown", 0.0)

    def _parse_fields(self, doc_type: str, lines: list[OCRLine], full_text: str, dedicated_mrz_lines: list = None) -> dict[str, FieldResult]:
        """Extract fields based on document type."""
        fields = {}

        if doc_type == "passport_rf":
            fields = self._parse_passport_rf(lines, full_text, dedicated_mrz_lines or [])
        elif doc_type == "passport_cis":
            fields = self._parse_passport_cis(lines, full_text)
        elif doc_type == "driver_license":
            fields = self._parse_dl(lines, full_text)
        elif doc_type == "snils":
            fields = self._parse_snils(lines, full_text)
        elif doc_type == "inn":
            fields = self._parse_inn(lines, full_text)
        elif doc_type == "bank_statement":
            # Bank statements are normally parsed directly from raw bytes
            # This case is for when OCR text happens to classify as bank_statement
            fields = {}

        return fields

    def _parse_passport_rf(self, lines, text, dedicated_mrz_lines=None) -> dict[str, FieldResult]:
        f = {}

        # ===== Strategy 1: MRZ parsing (загранпаспорт) =====
        # Cyrillic→Latin map for MRZ (EasyOCR reads MRZ letters as Cyrillic)
        # Standard Russian lookalikes: А В С Е Н К М О Р Т Х У З (no Ukrainian І)
        _cyr_to_lat_mrz = str.maketrans(
            "АВСЕНКМОРТХУЗавсенкмортхуз",
            "ABCEHKMOPTXUZABCEHKMOPTXUZ"
        )

        # Collect MRZ candidates from regular OCR
        mrz_lines = []
        for line in lines:
            clean = line.text.replace(" ", "")
            # Replace common OCR misreads of '<' filler
            clean = re.sub(r"[«»{}\[\]()~_\-=.,;:!?|/\\]", "<", clean)
            # Map Cyrillic lookalikes to Latin
            clean = clean.translate(_cyr_to_lat_mrz)
            # Remove remaining Cyrillic
            clean = re.sub(r"[а-яёА-ЯЁ]", "", clean)
            alpha_digits = sum(1 for c in clean if c.isalnum() or c == "<")
            if len(clean) >= 28 and alpha_digits / max(len(clean), 1) > 0.80:
                normalized = ""
                for c in clean.upper():
                    if c.isalpha() or c.isdigit() or c == "<":
                        normalized += c
                    else:
                        normalized += "<"
                # Smart K → < replacement (Tesseract often reads < as K)
                normalized = self._normalize_mrz_k_to_filler(normalized)
                if re.match(r"^[A-Z0-9<]+$", normalized):
                    mrz_lines.append(normalized)
                    logger.info(f"MRZ candidate (regular OCR): {normalized[:50]}...")

        # Merge dedicated MRZ lines with regular MRZ candidates
        if dedicated_mrz_lines:
            if len(dedicated_mrz_lines) >= 2:
                # Prefer dedicated if we got a full pair
                mrz_lines = dedicated_mrz_lines
                logger.info(f"Using dedicated MRZ lines ({len(mrz_lines)} found)")
            else:
                # Add dedicated lines to candidates
                mrz_lines.extend(dedicated_mrz_lines)
                logger.info(f"Added {len(dedicated_mrz_lines)} dedicated MRZ lines to {len(mrz_lines)} candidates")

        # Try to find the best MRZ pair: line starting with P and a digit line
        mrz_l1, mrz_l2 = None, None
        for ml in mrz_lines:
            if ml.startswith("P") and not mrz_l1:
                mrz_l1 = ml
            elif mrz_l1 and ml[0:1].isdigit():
                mrz_l2 = ml
                break

        # Fallback: just take the last two lines if we have >= 2
        if not (mrz_l1 and mrz_l2) and len(mrz_lines) >= 2:
            mrz_l1 = mrz_lines[-2]
            mrz_l2 = mrz_lines[-1]

        if mrz_l1 and mrz_l2:
            l1 = mrz_l1.ljust(44, "<")
            l2 = mrz_l2.ljust(44, "<")
            logger.info(f"MRZ L1 raw ({len(mrz_l1)} chars): '{mrz_l1}'")
            logger.info(f"MRZ L2 raw ({len(mrz_l2)} chars): '{mrz_l2}'")
            logger.info(f"MRZ L1 padded: '{l1[:44]}'")
            logger.info(f"MRZ L2 padded: '{l2[:44]}'")

            # Validate MRZ structure before parsing
            l1_valid = l1[0] == "P" and len(l1) >= 44
            l2_valid = l2[0:1].isdigit() and len(l2) >= 44
            logger.info(f"MRZ validation: L1_valid={l1_valid} (starts_with_P={l1[0]=='P'}), L2_valid={l2_valid}")

            try:
                # === Line 1: P<RUS SURNAME<<GIVENNAMES<<< ===
                country = l1[2:5].replace("<", "")
                name_part = l1[5:]
                logger.info(f"MRZ L1: country='{country}', name_part='{name_part[:30]}'")

                parts = name_part.split("<<", 1)
                ln = parts[0].replace("<", " ").strip()
                fn_raw = parts[1].replace("<", " ").strip() if len(parts) > 1 else ""
                fn_parts = fn_raw.split()
                logger.info(f"MRZ L1: surname='{ln}', given='{fn_raw}', fn_parts={fn_parts}")

                if ln and len(ln) >= 2:
                    f["last_name"] = FieldResult(ln, 0.95, "mrz")
                if fn_parts:
                    # Handle case where K remains as separator
                    if len(fn_parts) == 1 and "K" in fn_parts[0] and len(fn_parts[0]) > 6:
                        split_parts = re.split(r"(?<=[A-Z])K(?=[A-Z])", fn_parts[0])
                        if len(split_parts) >= 2 and all(len(p) >= 2 for p in split_parts):
                            fn_parts = split_parts
                            logger.info(f"MRZ: split given names on K separator: {fn_parts}")
                    f["first_name"] = FieldResult(fn_parts[0], 0.95, "mrz")
                    if len(fn_parts) > 1:
                        f["patronymic"] = FieldResult(" ".join(fn_parts[1:]), 0.90, "mrz")

                # === Line 2: fixed-position fields ===
                # TD3 format: [0-8]=doc_num, [9]=check, [10-12]=nationality,
                # [13-18]=DOB(YYMMDD), [19]=check, [20]=sex, [21-26]=expiry(YYMMDD), [27]=check
                logger.info(f"MRZ L2 positions: num='{l2[0:9]}' chk='{l2[9]}' nat='{l2[10:13]}' dob='{l2[13:19]}' chk='{l2[19]}' sex='{l2[20]}' exp='{l2[21:27]}' chk='{l2[27]}'")

                doc_num = l2[0:9].replace("<", "")
                if doc_num and len(doc_num) >= 7:
                    f["number"] = FieldResult(doc_num, 0.95, "mrz")
                    # Also extract series (first 2 digits) for Russian passports
                    if len(doc_num) >= 9:
                        f["series"] = FieldResult(doc_num[:2], 0.95, "mrz")

                nationality = l2[10:13].replace("<", "")

                bd = l2[13:19]
                if bd.isdigit() and len(bd) == 6:
                    yy = int(bd[:2])
                    yr = 1900 + yy if yy > 30 else 2000 + yy
                    f["birth_date"] = FieldResult(f"{bd[4:6]}.{bd[2:4]}.{yr}", 0.95, "mrz")
                    logger.info(f"MRZ DOB: raw='{bd}' → {f['birth_date'].value}")

                sex = l2[20]
                if sex in ("M", "F"):
                    f["sex"] = FieldResult("М" if sex == "M" else "Ж", 0.98, "mrz")
                    logger.info(f"MRZ sex: '{sex}' → {f['sex'].value}")

                exp = l2[21:27]
                if exp.isdigit() and len(exp) == 6:
                    eyy = int(exp[:2])
                    eyr = 2000 + eyy
                    f["expiry_date"] = FieldResult(f"{exp[4:6]}.{exp[2:4]}.{eyr}", 0.95, "mrz")
                    logger.info(f"MRZ expiry: raw='{exp}' → {f['expiry_date'].value}")

                logger.info(f"MRZ parsed fields: {list(f.keys())}")
                for k, v in f.items():
                    logger.info(f"  MRZ -> {k}: {v.value} ({v.confidence:.0%})")
            except Exception as e:
                logger.warning(f"MRZ parse error: {e}")

        # ===== Helper: clean OCR name text =====
        def clean_name_text(t):
            """Remove OCR garbage prefixes/suffixes from a name line.
            EasyOCR returns clean uppercase words, so preserve that.
            """
            if not t:
                return ""
            # If it's already uppercase Cyrillic (EasyOCR output), minimal processing
            if re.match(r"^[А-ЯЁ]+$", t.strip()):
                return t.strip()
            # Split into words and keep only words that are primarily Cyrillic
            words = t.split()
            clean_words = []
            for w in words:
                cyr_chars = sum(1 for c in w if '\u0400' <= c <= '\u04FF')
                total_alpha = sum(1 for c in w if c.isalpha())
                # Keep word if it's mostly Cyrillic
                # Short words (<=5 chars): need 100% Cyrillic. Longer words: 80%+
                min_ratio = 1.0 if total_alpha <= 5 else 0.8
                if total_alpha > 0 and cyr_chars / total_alpha >= min_ratio and cyr_chars >= 2:
                    # Strip non-Cyrillic chars from edges
                    cleaned_w = re.sub(r"^[^А-ЯЁа-яё]+", "", w)
                    cleaned_w = re.sub(r"[^А-ЯЁа-яё]+$", "", cleaned_w)
                    if cleaned_w and len(cleaned_w) >= 2:
                        clean_words.append(cleaned_w)
            if clean_words:
                return " ".join(clean_words)
            # Fallback: old cleaning method
            t = re.sub(r"^[^А-ЯЁа-яё]+", "", t)
            t = re.sub(r"[^А-ЯЁа-яё\'\-]+$", "", t)
            return t.strip()

        def is_valid_name(t):
            """Check if text looks like a Russian name (surname, first name, or patronymic).
            EasyOCR returns individual WORDS, so minimum valid length is 2 (e.g., 'ДЬ' in ДЬЯКОВ comes as separate WORD).
            """
            clean = t.strip().upper()
            if len(clean) < 2 or len(clean) > 25:
                return False
            cyr = sum(1 for c in clean if '\u0400' <= c <= '\u04FF')
            # For short words (2-3 chars), be lenient: just need Cyrillic
            # For longer words, need at least 2 Cyrillic chars (or proportional)
            if len(clean) <= 3:
                if cyr < 1:
                    return False
            else:
                if cyr < 2:
                    return False
            # Check ratio of Cyrillic chars
            alpha = sum(1 for c in clean if c.isalpha())
            if alpha == 0 or cyr / alpha < 0.7:
                return False
            # Reject known non-name patterns
            non_name = r"(?i)(паспорт|россий|россия|росаия|федера|граждан|район|област|город|улиц|посел|село|край|округ|республ|УФМС|МВД|ОВД|отдел|подразд|управлен|миграц|территор|пункт|выдач|ССР|штамп|печать|подпись|серия|номер|дата|место|рожден|жительств|фамил|имя|отчеств|личн|анчн|rижан|основн|регистр|орган|кол\b|код\b|russia|ritssia|passport|наманган|узбек|архангел|каргопол)"
            if re.search(non_name, clean):
                return False
            return True

        # Helper: extract Cyrillic part from bilingual line "ДЬЯКОВ/ DIAKOV"
        def extract_cyrillic(text_val):
            """Extract Cyrillic portion from bilingual text."""
            if "/" in text_val:
                parts = text_val.split("/")
                cyr = parts[0].strip()
                if cyr:
                    return cyr.upper()
            return text_val.strip().upper()

        # ===== Strategy 2: Visual zone OCR =====
        # Загранпаспорт has bilingual labels: "Фамилия / Surname" then "ДЬЯКОВ/ DIAKOV"

        # Helper: find value on same line or next line after label
        def find_value_after_label(label_patterns, lines_list, is_name_field=False):
            for i, line in enumerate(lines_list):
                for pat in label_patterns:
                    if re.search(pat, line.text, re.IGNORECASE):
                        # Check if value is on the SAME line (after label text)
                        remaining = re.sub(pat, "", line.text, flags=re.IGNORECASE).strip()
                        if remaining and len(remaining) >= 2 and not re.search(r"(?i)(surname|name|patronymic|birth|sex|place|дата|имя|фамилия|отчество|место|пол)", remaining):
                            return OCRLine(text=remaining, confidence=line.confidence, bbox=line.bbox, center_x=line.center_x, center_y=line.center_y)
                        for j in range(i + 1, min(i + 3, len(lines_list))):
                            next_line = lines_list[j].text.strip()
                            if re.search(r"(?i)(surname|name|patronymic|birth|sex|place|дата|имя|фамилия|отчество|место|пол|серия|номер|type|code|nationality|гражд|expiry|issue|орган|authority)", next_line):
                                continue
                            # For name fields, skip place-related text
                            if is_name_field and re.search(r"(?i)(район|област|город|улиц|посел|село|край|округ|республ|каргопол|архангел|проспект|деревн|ул\.|г\.|обл\.|пос\.)", next_line):
                                continue
                            if is_name_field and re.match(r"^[ВвНнПп]\s+[А-ЯЁа-яё]", next_line):
                                continue
                            if len(next_line) >= 2:
                                return lines_list[j]
            return None

        # Series + Number from visual zone
        if "number" not in f:
            for line in lines:
                lt = line.text
                # Internal passport: XX XX XXXXXX (e.g., 45 20 123456)
                m = re.search(r"(\d{2})\s+(\d{2})\s+(\d{6})", lt)
                if m:
                    f["series"] = FieldResult(f"{m.group(1)} {m.group(2)}", line.confidence, "ocr")
                    f["number"] = FieldResult(m.group(3), line.confidence, "ocr")
                    break
                # Загранпаспорт: 67N3817110 or 67 3817110
                m = re.search(r"(\d{2})\s*[NnНн№]?\s*(\d{7})", lt)
                if m:
                    f["number"] = FieldResult(f"{m.group(1)} {m.group(2)}", line.confidence, "ocr")
                    break
                # Compact 10 digits: XXXXXXXXXXXX (series 4 + number 6)
                t = lt.replace(" ", "").replace("O", "0").replace("o", "0")
                m = re.search(r"(\d{4})(\d{6})", t)
                if m:
                    f["series"] = FieldResult(f"{m.group(1)[:2]} {m.group(1)[2:]}", line.confidence, "ocr")
                    f["number"] = FieldResult(m.group(2), line.confidence, "ocr")
                    break

        # ===== Strategy C FIRST: Anchor-based (most reliable for internal passports) =====
        # On internal passports: Фамилия → ДЬЯКОВ → Имя → ЕГОР → Отчество → СЕРГЕЕВИЧ → Пол МУЖ. Дата 31.03.1988
        # The sex+date line is a reliable anchor; names are the 3 Cyrillic lines just above it
        if "last_name" not in f:
            anchor_idx = None
            anchor_y = None
            for i, line in enumerate(lines):
                # Look for a line with sex (МУЖ/ЖЕН) AND a date nearby
                if re.search(r"(?i)(муж|жен|мук|myx|myж)", line.text):
                    # Check if this line or adjacent lines have a date
                    # Flexible patterns: DD.MM.YYYY, DD,MM.YYYY, DD.MM YYYY, partial DD.MM + YYYY nearby
                    full_date = r"\d{2}[.,\-\s]\d{2}[.,\-\s]\d{4}"
                    partial_date = r"\d{2}[.,\-\s]\d{2}"
                    year_pat = r"\b(19[4-9]\d|20[0-2]\d)\b"
                    has_date = bool(re.search(full_date, line.text))
                    # Check partial date + year on same line or nearby
                    if not has_date and re.search(partial_date, line.text):
                        if re.search(year_pat, line.text):
                            has_date = True
                        else:
                            for di in range(-2, 3):
                                if 0 <= i + di < len(lines) and di != 0:
                                    if re.search(year_pat, lines[i + di].text):
                                        has_date = True
                                        break
                    if not has_date:
                        for di in range(-2, 3):
                            if 0 <= i + di < len(lines) and di != 0:
                                if re.search(full_date, lines[i + di].text):
                                    has_date = True
                                    break
                    # Last resort: МУЖ/ЖЕН alone is strong enough indicator (rare outside passport data page)
                    if not has_date:
                        has_date = True
                        logger.info(f"Strategy C: МУЖ/ЖЕН without date, using as weak anchor")
                    if has_date:
                        anchor_idx = i
                        anchor_y = line.center_y
                        logger.info(f"Strategy C anchor at line {i}: '{line.text}' y={line.center_y:.0f}")
                        break

            if anchor_idx is not None:
                # Collect name candidates: lines above anchor that look like names
                # On internal passport, names are within ~400px above the sex/date line
                name_candidates = []
                for j in range(len(lines)):
                    if j == anchor_idx:
                        continue
                    line = lines[j]
                    # Must be above the anchor
                    if line.center_y >= anchor_y - 10:
                        continue
                    # Must be within reasonable distance above anchor
                    if anchor_y - line.center_y > 400:
                        continue
                    # Clean the text and apply OCR corrections
                    raw = line.text.strip()
                    cleaned = clean_name_text(raw)
                    if cleaned:
                        cleaned = self._correct_ocr_names(cleaned)
                    if not cleaned:
                        continue
                    if not is_valid_name(cleaned):
                        continue
                    # Must have reasonable confidence
                    if line.confidence < 0.2:
                        continue
                    name_candidates.append((line, cleaned))
                    logger.info(f"Strategy C candidate: '{raw}' → '{cleaned}' y={line.center_y:.0f} conf={line.confidence:.0%}")

                # Sort by y position (top to bottom)
                name_candidates.sort(key=lambda x: x[0].center_y)

                # De-duplicate: if two candidates are very close vertically, keep the one with higher confidence
                deduped = []
                for line, cleaned in name_candidates:
                    skip = False
                    for existing_line, existing_cleaned in deduped:
                        if abs(line.center_y - existing_line.center_y) < 30:
                            # Same vertical position — keep higher confidence
                            if line.confidence > existing_line.confidence:
                                deduped.remove((existing_line, existing_cleaned))
                            else:
                                skip = True
                            break
                    if not skip:
                        deduped.append((line, cleaned))

                # Take the last 3 candidates (closest to anchor = surname, name, patronymic)
                final_names = deduped[-3:] if len(deduped) >= 3 else deduped
                names_order_c = ["last_name", "first_name", "patronymic"]
                for k, (nl, cleaned_val) in enumerate(final_names):
                    if k < len(names_order_c) and names_order_c[k] not in f:
                        val = cleaned_val.upper()
                        if len(val) >= 2:
                            f[names_order_c[k]] = FieldResult(val, nl.confidence * 0.80)
                            logger.info(f"Strategy C assigned {names_order_c[k]}: '{val}' (conf={nl.confidence:.0%})")

        # ===== Strategy D: Bilingual slash lines (загранпаспорт: "ДЬЯКОВ/ DIAKOV") =====
        # International passports have lines like "ДЬЯКОВ/" or "ДЬЯКОВ/ DIAKOV"
        if "last_name" not in f:
            bilingual_names = []
            for line in lines:
                lt = line.text.strip()
                # Pattern: Cyrillic/Latin or Cyrillic / Latin
                slash_m = re.match(r"^([А-ЯЁа-яё\s\'\-]+)\s*/\s*([A-Za-z\s\'\-]*)\s*$", lt)
                if slash_m:
                    cyr_part = slash_m.group(1).strip().upper()
                    cyr_part = self._correct_ocr_names(cyr_part)
                    if len(cyr_part) >= 2 and is_valid_name(cyr_part):
                        bilingual_names.append((cyr_part, line.confidence * 0.9, line.center_y))
                        logger.info(f"Strategy D bilingual: '{lt}' → Cyr='{cyr_part}'")
                # Also catch: just "ДЬЯКОВ/" at end of line
                elif lt.endswith("/") and re.match(r"^[А-ЯЁа-яё\s\'\-]+/$", lt):
                    cyr_part = lt[:-1].strip().upper()
                    cyr_part = self._correct_ocr_names(cyr_part)
                    if len(cyr_part) >= 2 and is_valid_name(cyr_part):
                        bilingual_names.append((cyr_part, line.confidence * 0.85, line.center_y))
                        logger.info(f"Strategy D trailing slash: '{lt}' → Cyr='{cyr_part}'")

            # Sort by Y position and assign: first = surname, second = name+patronymic
            bilingual_names.sort(key=lambda x: x[2])
            if bilingual_names:
                f["last_name"] = FieldResult(bilingual_names[0][0], bilingual_names[0][1])
                logger.info(f"Strategy D assigned last_name: '{bilingual_names[0][0]}'")
                if len(bilingual_names) >= 2:
                    name_text = bilingual_names[1][0]
                    parts = name_text.split()
                    if parts:
                        f["first_name"] = FieldResult(parts[0], bilingual_names[1][1])
                        if len(parts) > 1:
                            f["patronymic"] = FieldResult(" ".join(parts[1:]), bilingual_names[1][1] * 0.95)
                        logger.info(f"Strategy D assigned first_name: '{parts[0]}'")

        # ===== Strategy A: Label-based (for загранпаспорт with bilingual labels) =====
        if "last_name" not in f:
            val = find_value_after_label([r"(?i)фамил", r"(?i)surname"], lines, is_name_field=True)
            if val:
                cyr = extract_cyrillic(val.text)
                cleaned = clean_name_text(cyr)
                if cleaned:
                    cleaned = self._correct_ocr_names(cleaned)
                if cleaned and is_valid_name(cleaned):
                    f["last_name"] = FieldResult(cleaned.upper(), val.confidence * 0.85)
                    logger.info(f"Strategy A last_name: '{cleaned.upper()}'")

        if "first_name" not in f:
            val = find_value_after_label([r"(?i)\bимя\b", r"(?i)\bname\b", r"(?i)given"], lines, is_name_field=True)
            if val:
                cyr = extract_cyrillic(val.text)
                cleaned = clean_name_text(cyr)
                if cleaned:
                    cleaned = self._correct_ocr_names(cleaned)
                if cleaned and is_valid_name(cleaned):
                    parts = cleaned.split()
                    f["first_name"] = FieldResult(parts[0].upper(), val.confidence * 0.85)
                    if len(parts) > 1 and "patronymic" not in f:
                        f["patronymic"] = FieldResult(" ".join(parts[1:]).upper(), val.confidence * 0.80)

        if "patronymic" not in f:
            val = find_value_after_label([r"(?i)отчеств", r"(?i)patronym"], lines, is_name_field=True)
            if val:
                cyr = extract_cyrillic(val.text)
                cleaned = clean_name_text(cyr)
                if cleaned:
                    cleaned = self._correct_ocr_names(cleaned)
                if cleaned and is_valid_name(cleaned):
                    f["patronymic"] = FieldResult(cleaned.upper(), val.confidence * 0.8)

        # Strategy B: pure Cyrillic name lines (last resort fallback)
        # Only used if Strategies C and A didn't find names
        if "last_name" not in f:
            cyrillic = []
            for l in lines:
                raw = l.text.strip()
                cleaned = clean_name_text(raw)
                if cleaned:
                    cleaned = self._correct_ocr_names(cleaned)
                if not cleaned or not is_valid_name(cleaned):
                    continue
                # Must be single word (internal passport has 1 name per line)
                words = cleaned.split()
                if len(words) > 2:
                    continue
                # Must have minimum length and confidence
                if len(cleaned) < 3:
                    continue
                if l.confidence < 0.3:
                    continue
                cyrillic.append((l, cleaned))
                logger.info(f"Strategy B candidate: '{raw}' → '{cleaned}' conf={l.confidence:.0%} y={l.center_y:.0f}")

            # Sort by y position
            cyrillic.sort(key=lambda x: x[0].center_y)
            names_order = ["last_name", "first_name", "patronymic"]
            assigned = 0
            for line, cleaned_val in cyrillic:
                if assigned >= 3:
                    break
                if names_order[assigned] not in f:
                    f[names_order[assigned]] = FieldResult(cleaned_val.upper(), line.confidence * 0.85)
                    logger.info(f"Strategy B assigned {names_order[assigned]}: '{cleaned_val.upper()}'")
                    assigned += 1

        # Dates from visual zone
        all_dates = []
        for idx, line in enumerate(lines):
            for m in re.finditer(r"(\d{2})[.,\-\s](\d{2})[.,\-\s](\d{4})", line.text):
                date = f"{m.group(1)}.{m.group(2)}.{m.group(3)}"
                year = int(m.group(3))
                has_sex = bool(re.search(r"(?i)(муж|жен|мук|myx|myж)", line.text))
                has_issue = bool(re.search(r"(?i)(выдач|issue)", line.text))
                all_dates.append((date, year, line.confidence, line.text, has_sex, has_issue))

            # Also try to reconstruct split dates: "31,03" on one line + "1988" on next line
            partial_m = re.search(r"(\d{2})[.,\-](\d{2})$", line.text.strip())
            if not partial_m:
                partial_m = re.search(r"(\d{2})[.,\-](\d{2})\s*$", line.text)
            if partial_m:
                # Check next 2 lines for a year
                for di in range(1, 3):
                    if idx + di < len(lines):
                        year_m = re.match(r"^\s*(\d{4})\b", lines[idx + di].text.strip())
                        if year_m:
                            year = int(year_m.group(1))
                            date = f"{partial_m.group(1)}.{partial_m.group(2)}.{year}"
                            has_sex = bool(re.search(r"(?i)(муж|жен|мук|myx|myж)", line.text))
                            has_issue = bool(re.search(r"(?i)(выдач|issue)", line.text))
                            all_dates.append((date, year, line.confidence, line.text, has_sex, has_issue))
                            logger.info(f"Reconstructed split date: {date} from line '{line.text}' + '{lines[idx + di].text}'")
                            break

        # First pass: assign dates that have strong context clues (sex line = birth, issue label = issue)
        for date, year, conf, ctx, has_sex, has_issue in all_dates:
            if has_sex and 1940 <= year <= 2015 and "birth_date" not in f:
                f["birth_date"] = FieldResult(date, conf)
                logger.info(f"Birth date (sex-line context): {date}")
            elif has_issue and "issue_date" not in f:
                f["issue_date"] = FieldResult(date, conf)

        # Second pass: assign remaining dates by year range heuristics
        for date, year, conf, ctx, has_sex, has_issue in all_dates:
            if has_sex or has_issue:
                continue  # Already handled
            if 1940 <= year <= 2010 and "birth_date" not in f:
                f["birth_date"] = FieldResult(date, conf)
            elif 2000 <= year <= 2030 and "issue_date" not in f:
                f["issue_date"] = FieldResult(date, conf)
            elif 2025 <= year <= 2040 and "expiry_date" not in f:
                f["expiry_date"] = FieldResult(date, conf)

        # Series + number (internal passport: XX XX XXXXXX, international: XX XXXXXXX)
        if "number" not in f:
            # Internal passport: 11 07 574170
            m_ser = re.search(r"(\d{2})\s+(\d{2})\s+(\d{6})", text)
            if m_ser:
                f["series"] = FieldResult(f"{m_ser.group(1)} {m_ser.group(2)}", 0.9)
                f["number"] = FieldResult(m_ser.group(3), 0.9)
            else:
                # International passport: 67 3817110 (2 digits series + 7 digits number)
                m_int = re.search(r"(\d{2})\s+(\d{7})", text)
                if m_int:
                    f["series"] = FieldResult(m_int.group(1), 0.85)
                    f["number"] = FieldResult(m_int.group(2), 0.85)
                else:
                    # Compact 9 digits (international passport without spaces)
                    for line in lines:
                        digits = re.sub(r"\D", "", line.text)
                        if len(digits) == 9 and not re.search(r"\d{2}[./]\d{2}[./]\d{4}", line.text):
                            f["series"] = FieldResult(digits[:2], 0.75)
                            f["number"] = FieldResult(digits[2:], 0.75)
                            break

        # Department code / Authority
        m = re.search(r"(\d{3})\s*[-–.]\s*(\d{3})", text)
        if m:
            f["department_code"] = FieldResult(f"{m.group(1)}-{m.group(2)}", 0.9)

        # Authority (МВД / ФМС / УФМС / ОТДЕЛ УФМС)
        if "issuer" not in f:
            # Try to find full issuer text after label
            issuer_line = find_value_after_label([r"(?i)кем.*выдан", r"(?i)authority", r"(?i)орган.*выдав"], lines)
            if issuer_line and len(issuer_line.text.strip()) >= 5:
                f["issuer"] = FieldResult(issuer_line.text.strip(), issuer_line.confidence * 0.85)
            else:
                # Fallback: find МВД/ФМС/УФМС patterns
                m_auth = re.search(r"(?i)(ОТДЕЛ[А-Я\s]*УФМС[А-Я\s]*\d*|УФМС[А-Я\s]*\d*|ФМС\s*\d+|МВД\s*\d+|MVD\s*\d+)", text)
                if m_auth:
                    f["issuer"] = FieldResult(m_auth.group(0).strip().upper(), 0.85)

        # Birth place
        bp = find_value_after_label([r"(?i)место.*рожд", r"(?i)place.*birth"], lines)
        if bp and "birth_place" not in f:
            f["birth_place"] = FieldResult(bp.text.strip(), bp.confidence * 0.8)

        # Sex: support both Cyrillic and Latin OCR output, many OCR variants
        if "sex" not in f:
            sex_patterns_m = [r"(?i)\bмуж", r"\bMY[XЖжx]", r"\bMUZ", r"\bMUJ", r"\bMALE\b", r"\bM/M\b", r"\bM\s*/\s*M\b"]
            sex_patterns_f = [r"(?i)\bжен", r"\b[XЖж]EH", r"\bFEMALE\b", r"\bF/F\b", r"\bZEN\b"]
            for pat in sex_patterns_m:
                if re.search(pat, text):
                    f["sex"] = FieldResult("М", 0.9)
                    break
            if "sex" not in f:
                for pat in sex_patterns_f:
                    if re.search(pat, text):
                        f["sex"] = FieldResult("Ж", 0.9)
                        break

        # Birth place
        bp = find_value_after_label([r"(?i)место.*рожд", r"(?i)place.*birth", r"(?i)MECTO", r"(?i)PLACE"], lines)
        if bp and "birth_place" not in f:
            f["birth_place"] = FieldResult(bp.text.strip(), bp.confidence * 0.8)

        # === Fallback: extract dates from digit-only patterns ===
        # Look for dd.mm.yyyy patterns even without labels
        if "birth_date" not in f:
            for line in lines:
                for m in re.finditer(r"(\d{2})[.\-/\s](\d{2})[.\-/\s](\d{4})", line.text):
                    year = int(m.group(3))
                    if 1940 <= year <= 2015:
                        f["birth_date"] = FieldResult(f"{m.group(1)}.{m.group(2)}.{m.group(3)}", line.confidence * 0.8, "ocr")
                        break
                if "birth_date" in f:
                    break

        return f

    def _parse_passport_cis(self, lines, text) -> dict[str, FieldResult]:
        f = {}
        # MRZ
        mrz = [l.text.replace(" ", "") for l in lines if len(l.text.replace(" ", "")) >= 30 and re.match(r"^[A-Z0-9<]+$", l.text.replace(" ", ""))]
        if len(mrz) >= 2:
            l1, l2 = mrz[-2].ljust(44, "<"), mrz[-1].ljust(44, "<")
            try:
                country = l1[2:5].replace("<", "")
                parts = l1[5:].split("<<", 1)
                ln = parts[0].replace("<", " ").strip()
                fn = parts[1].replace("<", " ").strip() if len(parts) > 1 else ""
                num = l2[0:9].replace("<", "")
                bd = l2[13:19]
                sex = l2[20]

                if ln: f["last_name"] = FieldResult(ln, 0.95, "mrz")
                if fn:
                    ps = fn.split()
                    f["first_name"] = FieldResult(ps[0], 0.95, "mrz")
                    if len(ps) > 1: f["patronymic"] = FieldResult(" ".join(ps[1:]), 0.90, "mrz")
                if num: f["number"] = FieldResult(num, 0.95, "mrz")
                if country: f["country_code"] = FieldResult(country, 0.98, "mrz")
                if bd and bd.isdigit():
                    yy = int(bd[:2])
                    yr = 1900 + yy if yy > 30 else 2000 + yy
                    f["birth_date"] = FieldResult(f"{bd[4:6]}.{bd[2:4]}.{yr}", 0.95, "mrz")
                if sex in ("M", "F"):
                    f["sex"] = FieldResult("М" if sex == "M" else "Ж", 0.98, "mrz")
            except Exception:
                pass
        return f

    def _parse_dl(self, lines, text) -> dict[str, FieldResult]:
        """Parse Russian driving license (водительское удостоверение).
        Two-strategy approach:
        1. Label-based: find numbered fields (1., 2., 3., 4a., 4b., 5., etc.)
        2. Content-based: find data by patterns (names, dates, numbers)
        DL has bilingual lines: КИРИЛЛИЦА / LATIN or КИРИЛЛИЦА under Latin."""
        f = {}

        logger.info(f"DL parser: {len(lines)} lines")
        for i, l in enumerate(lines):
            logger.info(f"  DL line {i}: '{l.text}' (conf={l.confidence:.2f})")

        # ===== Strategy 1: Label-based (numbered fields on DL) =====
        # Russian DL fields:
        # 1. Фамилия / Surname
        # 2. Имя Отчество / Name Patronymic
        # 3. Дата рождения / Date of birth
        # 4a. Дата выдачи / Date of issue
        # 4b. Срок действия / Expiry date
        # 4c. Кем выдан / Issued by (ГИБДД code)
        # 5. Номер / Number
        # 6. Фото
        # 7. Подпись
        # 8. Место жительства / Residence
        # 9. Категории / Categories

        # First pass: try to find fields by label numbers in OCR text
        # Look in full text for patterns like "1. ДЬЯКОВА" or "1 ДЬЯКОВА"
        for i, line in enumerate(lines):
            lt = line.text.strip()

            # Field 1: Surname — "1. ДЬЯКОВА" or "1.ДЬЯКОВА / D'IAKOVA"
            m1 = re.match(r"^\s*1\s*[.)]\s*(.+)", lt)
            if m1 and "last_name" not in f:
                val = m1.group(1).strip()
                # Extract Cyrillic part (before "/" if bilingual)
                if "/" in val:
                    val = val.split("/")[0].strip()
                cyr_only = re.sub(r"[^А-ЯЁа-яё\s\'\-]", "", val).strip()
                if cyr_only and len(cyr_only) >= 2:
                    corrected = self._correct_ocr_names(cyr_only.upper())
                    f["last_name"] = FieldResult(corrected, line.confidence * 0.9)
                    logger.info(f"DL label 1 → last_name: '{corrected}'")
                continue

            # Field 2: Name + Patronymic — "2. ДИНА АНДРЕЕВНА" or "2.ДИНА АНДРЕЕВНА / DINA ANDREEVNA"
            m2 = re.match(r"^\s*2\s*[.)]\s*(.+)", lt)
            if m2 and "first_name" not in f:
                val = m2.group(1).strip()
                if "/" in val:
                    val = val.split("/")[0].strip()
                cyr_only = re.sub(r"[^А-ЯЁа-яё\s\'\-]", "", val).strip()
                if cyr_only:
                    corrected = self._correct_ocr_names(cyr_only.upper())
                    parts = corrected.split()
                    if parts:
                        f["first_name"] = FieldResult(parts[0], line.confidence * 0.9)
                        if len(parts) > 1:
                            f["patronymic"] = FieldResult(" ".join(parts[1:]), line.confidence * 0.85)
                        logger.info(f"DL label 2 → first_name: '{parts[0]}'" + (f", patronymic: '{' '.join(parts[1:])}'" if len(parts) > 1 else ""))
                continue

            # Field 3: Birth date — "3. 20.07.1995"
            m3 = re.match(r"^\s*3\s*[.)]\s*(.+)", lt)
            if m3 and "birth_date" not in f:
                val = m3.group(1).strip()
                dm = re.search(r"(\d{2})[.\-/](\d{2})[.\-/](\d{4})", val)
                if dm:
                    f["birth_date"] = FieldResult(f"{dm.group(1)}.{dm.group(2)}.{dm.group(3)}", line.confidence * 0.95)
                    logger.info(f"DL label 3 → birth_date: '{f['birth_date'].value}'")
                    # Also check for place after date on same line or label-like text
                    place_after = val[dm.end():].strip()
                    place_after = re.sub(r"^[,.\s]+", "", place_after).strip()
                    if place_after and len(place_after) >= 4:
                        f["birth_place"] = FieldResult(place_after.upper(), line.confidence * 0.8)
                continue

            # Field 4a: Issue date — "4a. 18.06.2024" or "4а) 18.06.2024"
            m4a = re.match(r"^\s*4\s*[аaа]\s*[.)]\s*(.+)", lt, re.IGNORECASE)
            if m4a and "issue_date" not in f:
                val = m4a.group(1).strip()
                dm = re.search(r"(\d{2})[.\-/](\d{2})[.\-/](\d{4})", val)
                if dm:
                    f["issue_date"] = FieldResult(f"{dm.group(1)}.{dm.group(2)}.{dm.group(3)}", line.confidence * 0.95)
                    logger.info(f"DL label 4a → issue_date: '{f['issue_date'].value}'")
                    # Check if 4b is on same line: "4a) 18.06.2024  4b) 18.06.2034"
                    rest = val[dm.end():]
                    m4b_same = re.search(r"4\s*[bб]\s*[.)]\s*(\d{2})[.\-/](\d{2})[.\-/](\d{4})", rest, re.IGNORECASE)
                    if m4b_same and "expiry_date" not in f:
                        f["expiry_date"] = FieldResult(f"{m4b_same.group(1)}.{m4b_same.group(2)}.{m4b_same.group(3)}", line.confidence * 0.95)
                        logger.info(f"DL label 4b (same line) → expiry_date: '{f['expiry_date'].value}'")
                continue

            # Field 4b: Expiry date — "4b. 18.06.2034" or "4б) 18.06.2034"
            m4b = re.match(r"^\s*4\s*[bб]\s*[.)]\s*(.+)", lt, re.IGNORECASE)
            if m4b and "expiry_date" not in f:
                val = m4b.group(1).strip()
                dm = re.search(r"(\d{2})[.\-/](\d{2})[.\-/](\d{4})", val)
                if dm:
                    f["expiry_date"] = FieldResult(f"{dm.group(1)}.{dm.group(2)}.{dm.group(3)}", line.confidence * 0.95)
                    logger.info(f"DL label 4b → expiry_date: '{f['expiry_date'].value}'")
                continue

            # Field 4c: Issued by ГИБДД — "4c. ГИБДД 2301" or "4с) ГИБДД2301"
            m4c = re.match(r"^\s*4\s*[cсc]\s*[.)]\s*(.+)", lt, re.IGNORECASE)
            if m4c and "issuer" not in f:
                val = m4c.group(1).strip()
                m_gibdd = re.search(r"(?i)(ГИБДД|GIBDD|ГИ6ДД)\s*(\d{2,5})", val)
                if m_gibdd:
                    f["issuer"] = FieldResult(f"ГИБДД {m_gibdd.group(2)}", line.confidence * 0.9)
                elif val:
                    f["issuer"] = FieldResult(val.upper(), line.confidence * 0.8)
                logger.info(f"DL label 4c → issuer: '{f.get('issuer', {})}'")
                continue

            # Field 5: DL number — "5. 99 33 750095"
            m5 = re.match(r"^\s*5\s*[.)]\s*(.+)", lt)
            if m5 and "number" not in f:
                val = m5.group(1).strip()
                mn = re.search(r"(\d{2})\s*(\d{2})\s*(\d{6})", val)
                if mn:
                    f["number"] = FieldResult(f"{mn.group(1)} {mn.group(2)} {mn.group(3)}", line.confidence * 0.95)
                    logger.info(f"DL label 5 → number: '{f['number'].value}'")
                continue

            # Field 8: Residence — "8. АРХАНГЕЛЬСКАЯ ОБЛ."
            m8 = re.match(r"^\s*8\s*[.)]\s*(.+)", lt)
            if m8:
                val = m8.group(1).strip()
                if len(val) >= 4 and "residence" not in f:
                    f["residence"] = FieldResult(val.upper(), line.confidence * 0.8)
                    logger.info(f"DL label 8 → residence: '{val.upper()}'")
                continue

        # ===== Strategy 2: Content-based fallback =====
        skip_words = r"(?i)(водительск|удостовер|permis|conduire|driving|licen[cs]e|российск|федерац|RUS\b)"
        name_lines_cyr = []
        name_lines_lat = []
        date_lines = []
        number_line = None

        for i, line in enumerate(lines):
            lt = line.text.strip()
            clean = re.sub(r"^\s*\d+[.)а-яa-z]*\s*[.)]\s*", "", lt).strip()
            if not clean:
                clean = lt

            if re.search(skip_words, clean):
                continue

            # === DL number: XX XX XXXXXX ===
            if "number" not in f:
                m = re.search(r"(\d{2})\s+(\d{2})\s+(\d{6})", clean)
                if m:
                    number_line = (f"{m.group(1)} {m.group(2)} {m.group(3)}", line.confidence)
                    continue
                digits_only = re.sub(r"\D", "", clean)
                if len(digits_only) == 10 and not re.search(r"\d{2}[./]\d{2}[./]\d{4}", clean):
                    number_line = (f"{digits_only[:2]} {digits_only[2:4]} {digits_only[4:]}", line.confidence * 0.85)
                    continue

            # === Dates ===
            date_matches = re.findall(r"(\d{2})[.\-/](\d{2})[.\-/](\d{4})", clean)
            if date_matches:
                for dm in date_matches:
                    date_str = f"{dm[0]}.{dm[1]}.{dm[2]}"
                    year = int(dm[2])
                    date_lines.append((date_str, year, line.confidence, clean))
                continue

            # === ГИБДД ===
            if "issuer" not in f:
                m_gibdd = re.search(r"(?i)(ГИБДД|GIBDD|ГИ6ДД|ГИБД)\s*(\d{2,4})", clean)
                if m_gibdd:
                    f["issuer"] = FieldResult(f"ГИБДД {m_gibdd.group(2)}", 0.9)
                    continue

            # === Place names ===
            if re.search(r"(?i)(обл\.?|край|респ|МОСКВ|ПЕТЕРБ|ОБЛ\b|OBL|KRAI|город)", clean) and len(clean) >= 5:
                place_text = re.sub(r"\d{2}[.\-/]\d{2}[.\-/]\d{4}", "", clean).strip()
                place_text = re.sub(r"^\s*\d+\s*[.)]\s*", "", place_text).strip()
                if place_text and len(place_text) >= 4:
                    # Try to prefer Cyrillic text
                    cyr_match = re.search(r"[А-ЯЁа-яё]{3,}", place_text)
                    if cyr_match:
                        place_text = cyr_match.group(0)
                    if "birth_place" not in f:
                        f["birth_place"] = FieldResult(place_text.upper(), line.confidence * 0.8)
                    elif "residence" not in f:
                        f["residence"] = FieldResult(place_text.upper(), line.confidence * 0.8)
                continue

            # === Bilingual name lines: "ДЬЯКОВА / D'IAKOVA" ===
            if "/" in clean:
                parts = clean.split("/")
                cyr_part = parts[0].strip()
                lat_part = parts[1].strip() if len(parts) > 1 else ""
                if re.match(r"^[А-ЯЁа-яё\s\'\-]{2,40}$", cyr_part):
                    corrected_cyr = self._correct_ocr_names(cyr_part.upper())
                    name_lines_cyr.append((corrected_cyr, line.confidence * 0.9))
                    if lat_part:
                        name_lines_lat.append((lat_part.upper(), line.confidence * 0.8))
                    continue

            # === Pure Cyrillic name ===
            if re.match(r"^[А-ЯЁа-яё\s\'\-]{2,40}$", clean) and len(clean) >= 3:
                corrected_clean = self._correct_ocr_names(clean.upper())
                name_lines_cyr.append((corrected_clean, line.confidence * 0.85))
                continue

            # === Pure Latin name ===
            if re.match(r"^[A-Z][A-Z\'\s\-]{1,35}$", clean) and len(clean) >= 3:
                if not re.search(r"(?i)(RUS|PERMIS|DRIVING|LICEN|CONDUIRE|GIBDD)", clean):
                    name_lines_lat.append((clean.upper(), line.confidence * 0.8))
                continue

        # === Assign FIO from content-based name lines (if label-based didn't find) ===
        logger.info(f"DL name candidates (Cyrillic): {name_lines_cyr}")
        logger.info(f"DL name candidates (Latin): {name_lines_lat}")

        name_lines_cyr = [(t, c) for t, c in name_lines_cyr if len(t.replace(" ", "")) >= 3]

        if "last_name" not in f and len(name_lines_cyr) >= 1:
            first_name_text, first_conf = name_lines_cyr[0]
            first_words = first_name_text.split()

            if len(first_words) == 1:
                f["last_name"] = FieldResult(first_words[0], first_conf)
                if len(name_lines_cyr) >= 2:
                    second_text, second_conf = name_lines_cyr[1]
                    parts = second_text.split()
                    if parts:
                        f["first_name"] = FieldResult(parts[0], second_conf)
                        if len(parts) > 1:
                            f["patronymic"] = FieldResult(" ".join(parts[1:]), second_conf * 0.95)
            elif len(first_words) >= 2:
                if len(name_lines_cyr) >= 2:
                    second_text, second_conf = name_lines_cyr[1]
                    second_words = second_text.split()
                    if len(second_words) == 1:
                        f["last_name"] = FieldResult(second_words[0], second_conf)
                        f["first_name"] = FieldResult(first_words[0], first_conf)
                        if len(first_words) > 1:
                            f["patronymic"] = FieldResult(" ".join(first_words[1:]), first_conf * 0.95)
                    else:
                        f["last_name"] = FieldResult(first_words[0], first_conf)
                        f["first_name"] = FieldResult(second_words[0], second_conf)
                        if len(second_words) > 1:
                            f["patronymic"] = FieldResult(" ".join(second_words[1:]), second_conf * 0.95)
                else:
                    f["last_name"] = FieldResult(first_words[0], first_conf)
                    if len(first_words) > 1:
                        f["first_name"] = FieldResult(first_words[1], first_conf * 0.9)
                    if len(first_words) > 2:
                        f["patronymic"] = FieldResult(" ".join(first_words[2:]), first_conf * 0.85)

        # === Latin→Cyrillic transliteration helper (must be defined before use) ===
        _lat_to_cyr = {
            "A": "А", "B": "Б", "V": "В", "G": "Г", "D": "Д", "E": "Е", "ZH": "Ж",
            "Z": "З", "I": "И", "Y": "Й", "K": "К", "L": "Л", "M": "М", "N": "Н",
            "O": "О", "P": "П", "R": "Р", "S": "С", "T": "Т", "U": "У", "F": "Ф",
            "KH": "Х", "TS": "Ц", "CH": "Ч", "SH": "Ш", "SHCH": "Щ",
            "IU": "Ю", "IA": "Я", "YU": "Ю", "YA": "Я",
            "EV": "ЕВ", "EI": "ЕЙ",
        }
        def _transliterate_to_cyr(lat_name):
            """Reverse-transliterate Latin name to Cyrillic (best-effort)."""
            result = ""
            s = lat_name.upper().replace("'", "Ь")
            i = 0
            while i < len(s):
                # Try 4-char, 3-char, 2-char, 1-char matches
                matched = False
                for length in [4, 3, 2]:
                    chunk = s[i:i+length]
                    if chunk in _lat_to_cyr:
                        result += _lat_to_cyr[chunk]
                        i += length
                        matched = True
                        break
                if not matched:
                    if s[i] in _lat_to_cyr:
                        result += _lat_to_cyr[s[i]]
                    elif s[i] == "Ь":
                        result += "Ь"
                    else:
                        result += s[i]  # Keep as-is
                    i += 1
            return result

        # === Fix: Override low-confidence first_name with Latin transliteration ===
        if "first_name" in f and f["first_name"].confidence < 0.50 and name_lines_lat:
            valid_lat_fn = [(t, c) for t, c in name_lines_lat if len(t.replace(" ", "")) >= 3 and c >= 0.70]
            if valid_lat_fn:
                lat_name = valid_lat_fn[0][0]
                lat_conf = valid_lat_fn[0][1]
                if lat_conf > f["first_name"].confidence:
                    cyr_name = _transliterate_to_cyr(lat_name.split()[0] if ' ' in lat_name else lat_name)
                    if cyr_name and len(cyr_name) >= 2:
                        logger.info(f"DL first_name override: '{f['first_name'].value}' → Latin '{lat_name}' → '{cyr_name}'")
                        f["first_name"] = FieldResult(cyr_name, lat_conf)

        use_latin_fallback = False
        if "last_name" not in f:
            use_latin_fallback = True
        elif f.get("last_name") and f["last_name"].confidence < 0.5:
            use_latin_fallback = True

        if use_latin_fallback and name_lines_lat:
            valid_lat = [(t, c) for t, c in name_lines_lat if len(t.replace(" ", "")) >= 3]
            if valid_lat:
                cyr_name = _transliterate_to_cyr(valid_lat[0][0])
                if cyr_name and len(cyr_name) >= 2:
                    f["last_name"] = FieldResult(cyr_name, valid_lat[0][1])
                    logger.info(f"DL Latin fallback last_name: '{valid_lat[0][0]}' → '{cyr_name}'")
                if len(valid_lat) >= 2:
                    parts = valid_lat[1][0].split()
                    if parts:
                        cyr_fn = _transliterate_to_cyr(parts[0])
                        f["first_name"] = FieldResult(cyr_fn, valid_lat[1][1])
                        logger.info(f"DL Latin fallback first_name: '{parts[0]}' → '{cyr_fn}'")
                        if len(parts) > 1:
                            cyr_patr = _transliterate_to_cyr(" ".join(parts[1:]))
                            f["patronymic"] = FieldResult(cyr_patr, valid_lat[1][1] * 0.95)
                            logger.info(f"DL Latin fallback patronymic: '{' '.join(parts[1:])}' → '{cyr_patr}'")

        # === Assign dates ===
        date_lines.sort(key=lambda x: x[1])  # Sort by year
        logger.info(f"DL dates found: {[(d, y) for d, y, _, _ in date_lines]}")

        for date_str, year, conf, ctx in date_lines:
            day = int(date_str[:2]) if date_str[:2].isdigit() else 0
            month = int(date_str[3:5]) if date_str[3:5].isdigit() else 0
            if not (1 <= day <= 31 and 1 <= month <= 12):
                logger.info(f"DL skipping invalid date: '{date_str}'")
                continue
            if 1940 <= year <= 2012 and "birth_date" not in f:
                f["birth_date"] = FieldResult(date_str, conf)
            elif 2010 <= year <= 2026 and "issue_date" not in f:
                f["issue_date"] = FieldResult(date_str, conf)
            elif year > 2026 and "expiry_date" not in f:
                f["expiry_date"] = FieldResult(date_str, conf)

        # Fallback: two future dates → earlier = issue, later = expiry
        if ("issue_date" not in f or "expiry_date" not in f):
            future = [(d, y, c) for d, y, c, _ in date_lines if y >= 2020]
            future.sort(key=lambda x: x[1])
            if len(future) >= 2:
                if "issue_date" not in f:
                    f["issue_date"] = FieldResult(future[0][0], future[0][2])
                if "expiry_date" not in f:
                    f["expiry_date"] = FieldResult(future[-1][0], future[-1][2])

        # === Assign number ===
        if number_line:
            f["number"] = FieldResult(number_line[0], number_line[1])

        # === Categories ===
        # On Russian DL, categories are in boxes at bottom: "B", "B1", "M" (may be separate lines)
        # IMPORTANT: Avoid false positives like "D" from "DRIVING" text
        dl_categories = set()
        valid_cats = {"A", "A1", "B", "B1", "BE", "C", "C1", "CE", "C1E", "D", "D1", "DE", "D1E", "M", "Tm", "Tb"}
        for line in lines:
            lt = line.text.strip()

            # Skip lines that contain common OCR mistakes or text that shouldn't have categories
            if re.search(r"(?i)(driving|licence|licen|permis|conduire|водит|удосто)", lt):
                continue

            # Category line: short, contains only valid category letters
            cats = re.findall(r"\b([ABCDEM][12E]?)\b", lt.upper())
            # Also catch single-char categories like "B" or "M" on very short lines
            if len(lt) <= 4 and lt.upper() in valid_cats:
                dl_categories.add(lt.upper())
                continue
            valid_found = set(cats).intersection(valid_cats)
            # "9." label or "категории"
            if re.match(r"^\s*9\s*[.)]\s*", lt) or re.search(r"(?i)категор|categor", lt):
                dl_categories.update(valid_found)
            elif len(valid_found) >= 2 and len(lt) <= 30:
                dl_categories.update(valid_found)
            elif len(valid_found) >= 1 and len(lt) <= 8:
                # Very short line with a category (e.g., "B1", "M", "B B1 M")
                # Only add if it looks like a category line (mostly letters and spaces/pipes)
                if re.match(r"^[\s|BbАВСDEМмMm12,.\[\]]*[ABCDEM][\s|BbАВСDEМмMm12,.\[\]]*$", lt.upper()):
                    dl_categories.update(valid_found)
            elif re.match(r"^[\s|BbАВСDEМмMm12,.\[\]]+$", lt) and len(lt) <= 25:
                dl_categories.update(valid_found)
            # Also try to find categories in box-like patterns: "[B]" or "|B|"
            box_cats = re.findall(r"[\[\|]([ABCDEM][12E]?)[\]\|]", lt.upper())
            dl_categories.update(set(box_cats).intersection(valid_cats))

        if dl_categories:
            cat_order = ["A", "A1", "B", "B1", "BE", "C", "C1", "CE", "C1E", "D", "D1", "DE", "D1E", "M"]
            sorted_cats = sorted(dl_categories, key=lambda c: cat_order.index(c) if c in cat_order else 99)
            f["categories"] = FieldResult(", ".join(sorted_cats), 0.85)
            logger.info(f"DL categories found: {sorted_cats}")

        logger.info(f"DL parsed fields: {list(f.keys())}")
        for k, v in f.items():
            logger.info(f"  DL -> {k}: {v.value} ({v.confidence:.0%})")

        return f

    def _parse_snils(self, lines, text) -> dict[str, FieldResult]:
        """Parse SNILS (Social Insurance Number) with checksum validation.
        Format: XXX-XXX-XXX XX (11 digits total with check digits)"""
        f = {}

        # Try to find SNILS pattern in text: 123-456-789 01
        m = re.search(r"(\d{3})[-\s]?(\d{3})[-\s]?(\d{3})[-\s]?(\d{2})", text.replace(" ", ""))
        if not m:
            # Fall back to searching individual lines
            for line in lines:
                digits = re.sub(r"\D", "", line.text)
                if len(digits) == 11:
                    m = re.match(r"(\d{3})(\d{3})(\d{3})(\d{2})", digits)
                    if m:
                        break

        if m:
            snils_num = f"{m.group(1)}{m.group(2)}{m.group(3)}{m.group(4)}"
            snils_formatted = f"{m.group(1)}-{m.group(2)}-{m.group(3)} {m.group(4)}"

            # Validate checksum
            num_part = snils_num[:9]
            check_digits = int(snils_num[9:11])
            total = sum(int(d) * (9 - i) for i, d in enumerate(num_part))

            if total < 100:
                expected = total
            elif total in (100, 101):
                expected = 0
            else:
                r = total % 101
                expected = 0 if r >= 100 else r

            is_valid = check_digits == expected
            confidence = 0.95 if is_valid else 0.70
            needs_review = not is_valid

            f["snils_number"] = FieldResult(
                snils_formatted,
                confidence,
                "ocr",
                needs_review=needs_review
            )
            logger.info(f"SNILS parsed: {snils_formatted} (checksum {'valid' if is_valid else 'INVALID'})")

        return f

    def _parse_inn(self, lines, text) -> dict[str, FieldResult]:
        """Parse INN (Tax Identification Number) with checksum validation.
        10 digits (legal entity) or 12 digits (individual)"""
        f = {}

        # Try to find INN pattern: 10 or 12 digit number
        m = re.search(r"\b(\d{10}|\d{12})\b", text)

        if m:
            inn_number = m.group(1)
            inn_type = "personal" if len(inn_number) == 12 else "legal"

            # Validate checksum
            is_valid = self._validate_inn_checksum(inn_number)
            confidence = 0.95 if is_valid else 0.70
            needs_review = not is_valid

            f["inn_number"] = FieldResult(
                inn_number,
                confidence,
                "ocr",
                needs_review=needs_review
            )
            f["inn_type"] = FieldResult(inn_type, 0.99, "ocr")
            logger.info(f"INN parsed: {inn_number} ({inn_type}, checksum {'valid' if is_valid else 'INVALID'})")

        return f

    def _validate_inn_checksum(self, inn_str: str) -> bool:
        """Validate INN checksum.
        10 digits (legal): one check digit
        12 digits (personal): two check digits"""
        if not inn_str or not inn_str.isdigit():
            return False

        if len(inn_str) == 10:
            # Legal entity: positions weighted [2,4,10,3,5,9,7,6,8]
            weights = [2, 4, 10, 3, 5, 9, 7, 6, 8]
            total = sum(int(inn_str[i]) * weights[i] for i in range(9))
            check = total % 11
            if check >= 10:
                check = 0
            return int(inn_str[9]) == check

        elif len(inn_str) == 12:
            # Individual: two check digits
            # First check digit (position 10)
            weights1 = [7, 2, 4, 10, 3, 5, 9, 7, 6, 8]
            total1 = sum(int(inn_str[i]) * weights1[i] for i in range(10))
            check1 = total1 % 11
            if check1 >= 10:
                check1 = 0

            if int(inn_str[10]) != check1:
                return False

            # Second check digit (position 11)
            weights2 = [3, 7, 2, 4, 10, 3, 5, 9, 7, 6, 8]
            total2 = sum(int(inn_str[i]) * weights2[i] for i in range(11))
            check2 = total2 % 11
            if check2 >= 10:
                check2 = 0

            return int(inn_str[11]) == check2

        return False

    def _parse_bank_statement(self, raw_bytes: bytes, fmt: str, fmt_data: dict) -> dict[str, FieldResult]:
        """Parse bank statement from TXT (1C format), CSV, or XLSX.
        Returns structured fields with high confidence (1.0) since this is not OCR."""
        f = {}

        try:
            if fmt == 'txt_1c':
                f = self._parse_bank_statement_txt_1c(raw_bytes, fmt_data)
            elif fmt == 'csv':
                f = self._parse_bank_statement_csv(raw_bytes, fmt_data)
            elif fmt == 'xlsx':
                f = self._parse_bank_statement_xlsx(raw_bytes, fmt_data)
            else:
                logger.warning(f"Unknown bank statement format: {fmt}")
                return {}

            logger.info(f"Bank statement parsed: {len(f)} fields extracted")
            return f

        except Exception as e:
            logger.error(f"Bank statement parsing error: {e}")
            raise

    def _parse_bank_statement_txt_1c(self, raw_bytes: bytes, fmt_data: dict) -> dict[str, FieldResult]:
        """Parse 1CClientBankExchange TXT format (Windows-1251 encoded)."""
        f = {}
        encoding = fmt_data.get('encoding', 'windows-1251')

        try:
            text = raw_bytes.decode(encoding, errors='ignore')
        except Exception:
            text = raw_bytes.decode('utf-8', errors='ignore')

        lines = text.split('\n')
        data = {}
        transactions = []
        current_doc = None

        for line in lines:
            line = line.strip()
            if not line:
                continue

            # Handle section markers WITHOUT '=' sign
            if line == 'СекцияРасчСчет':
                data['has_account_section'] = True
                continue
            elif line == 'КонецРасчСчет':
                continue
            elif line == 'КонецДокумента':
                if current_doc:
                    transactions.append(current_doc)
                current_doc = None
                continue

            if '=' not in line:
                continue

            key, _, value = line.partition('=')
            key = key.strip()
            value = value.strip()

            # Section markers WITH '=' sign
            if key == 'СекцияДокумент':
                current_doc = {'type': value}
                continue

            # Account section fields
            if key == 'РасчСчет':
                data['account_number'] = value
            elif key == 'НачальныйОстаток':
                data['opening_balance'] = value
            elif key == 'КонечныйОстаток':
                data['closing_balance'] = value
            elif key == 'ВсегоПоступило':
                data['total_credit'] = value
            elif key == 'ВсегоСписано':
                data['total_debit'] = value
            elif key == 'ДатаСоздания':
                data['created_date'] = value
            elif key == 'ДатаНачала':
                data['period_start'] = value
            elif key == 'ДатаКонца':
                data['period_end'] = value

            # Transaction fields
            if current_doc:
                if key == 'Номер':
                    current_doc['number'] = value
                elif key == 'Дата':
                    current_doc['date'] = value
                elif key == 'Сумма':
                    current_doc['amount'] = value
                elif key == 'ПлательщикСчет':
                    current_doc['payer_account'] = value
                elif key == 'ПлательщикИНН':
                    current_doc['payer_inn'] = value
                elif key == 'Плательщик1':
                    current_doc['payer_name'] = value
                elif key == 'ПолучательСчет':
                    current_doc['recipient_account'] = value
                elif key == 'ПолучательИНН':
                    current_doc['recipient_inn'] = value
                elif key == 'Получатель1':
                    current_doc['recipient_name'] = value
                elif key == 'НазначениеПлатежа':
                    current_doc['purpose'] = value
                elif key == 'ПолучательБИК':
                    current_doc['recipient_bik'] = value
                elif key == 'ПолучательКорсчет':
                    current_doc['recipient_corr_account'] = value
                elif key == 'ПолучательБанк1':
                    current_doc['recipient_bank'] = value
                elif key == 'ПлательщикБИК':
                    current_doc['payer_bik'] = value
                elif key == 'ПлательщикКорсчет':
                    current_doc['payer_corr_account'] = value
                elif key == 'ПлательщикБанк1':
                    current_doc['payer_bank'] = value

        # Determine our account and extract holder info
        our_account = data.get('account_number', '')
        for t in transactions:
            if t.get('payer_name') and t.get('payer_account') == our_account:
                data['account_holder'] = t['payer_name']
                break

        # Filter credit transactions only (incoming: recipient is us, payer is counterparty)
        credit_transactions = []
        credit_total = 0.0
        for t in transactions:
            # Credit = money coming IN: recipient_account == our account
            if t.get('recipient_account') == our_account:
                try:
                    amount = float((t.get('amount', '0') or '0').replace(',', '.').replace(' ', ''))
                except (ValueError, TypeError):
                    amount = 0.0
                if amount > 0:
                    credit_total += amount
                    txn = {
                        'bik': t.get('payer_bik', ''),
                        'corr_account': t.get('payer_corr_account', ''),
                        'bank_name': t.get('payer_bank', ''),
                        'counterparty_account': t.get('payer_account', ''),
                        'counterparty': t.get('payer_name', ''),
                        'doc_number': t.get('number', ''),
                        'doc_date': t.get('date', ''),
                        'credit': f'{amount:,.2f}',
                        'purpose': t.get('purpose', ''),
                        'counterparty_inn': t.get('payer_inn', ''),
                    }
                    credit_transactions.append(txn)

        f['account_number'] = FieldResult(our_account, 1.0, 'bank_statement_parsing')
        f['account_holder'] = FieldResult(data.get('account_holder', ''), 1.0, 'bank_statement_parsing')
        f['period_start'] = FieldResult(data.get('period_start', ''), 1.0, 'bank_statement_parsing')
        f['period_end'] = FieldResult(data.get('period_end', ''), 1.0, 'bank_statement_parsing')
        f['total_credit'] = FieldResult(f'{credit_total:,.2f} ₽', 1.0, 'bank_statement_parsing')
        f['transaction_count'] = FieldResult(str(len(credit_transactions)), 1.0, 'bank_statement_parsing')
        f['currency'] = FieldResult('RUB', 1.0, 'bank_statement_parsing')
        f['transactions'] = FieldResult(credit_transactions, 1.0, 'bank_statement_parsing')

        return f

    def _parse_bank_statement_csv(self, raw_bytes: bytes, fmt_data: dict) -> dict[str, FieldResult]:
        """Parse bank statement in CSV format (semicolon-separated, Windows-1251).
        Extracts only credit transactions (Кредит > 0) with specific columns."""
        f = {}
        encoding = fmt_data.get('encoding', 'windows-1251')

        try:
            text = raw_bytes.decode(encoding, errors='ignore')
        except Exception:
            text = raw_bytes.decode('utf-8', errors='ignore')

        # Parse CSV with semicolon delimiter
        reader = csv.DictReader(StringIO(text), delimiter=';')
        rows = list(reader)

        if not rows:
            return f

        first_row = rows[0]
        account = first_row.get('Счет', '')
        f['account_number'] = FieldResult(account, 1.0, 'bank_statement_parsing')
        f['account_holder'] = FieldResult('', 1.0, 'bank_statement_parsing')

        # Collect dates and build credit transactions
        dates = [row.get('Дата операции', '') for row in rows if row.get('Дата операции')]
        if dates:
            f['period_start'] = FieldResult(min(dates), 1.0, 'bank_statement_parsing')
            f['period_end'] = FieldResult(max(dates), 1.0, 'bank_statement_parsing')

        credit_total = 0.0
        credit_transactions = []
        for row in rows:
            try:
                credit_str = (row.get('Кредит', '') or '').replace(',', '.').replace(' ', '')
                credit = float(credit_str) if credit_str else 0.0
                if credit > 0:
                    credit_total += credit
                    txn = {
                        'bik': row.get('Бик банка', ''),
                        'corr_account': row.get('Кор. счет банка', ''),
                        'bank_name': row.get('Название банка', ''),
                        'counterparty_account': row.get('Счет корреспондента', ''),
                        'counterparty': row.get('Корреспондент', ''),
                        'doc_number': row.get('Номер документа', ''),
                        'doc_date': row.get('Дата документа', ''),
                        'credit': f'{credit:,.2f}',
                        'purpose': row.get('Примечание', ''),
                        'counterparty_inn': row.get('ИНН Корреспондента', ''),
                    }
                    credit_transactions.append(txn)
            except (ValueError, TypeError):
                pass

        f['total_credit'] = FieldResult(f'{credit_total:,.2f} ₽', 1.0, 'bank_statement_parsing')
        f['transaction_count'] = FieldResult(str(len(credit_transactions)), 1.0, 'bank_statement_parsing')
        f['currency'] = FieldResult('RUB', 1.0, 'bank_statement_parsing')
        f['transactions'] = FieldResult(credit_transactions, 1.0, 'bank_statement_parsing')

        return f

    def _parse_bank_statement_xlsx(self, raw_bytes: bytes, fmt_data: dict) -> dict[str, FieldResult]:
        """Parse bank statement in XLSX format."""
        f = {}

        if not HAS_OPENPYXL:
            logger.warning("openpyxl not installed, cannot parse XLSX")
            return f

        try:
            from io import BytesIO
            xlsx_file = BytesIO(raw_bytes)
            wb = openpyxl.load_workbook(xlsx_file)

            if 'Выписка' not in wb.sheetnames:
                logger.warning("No 'Выписка' sheet found in XLSX")
                return f

            ws = wb['Выписка']
            all_rows = list(ws.iter_rows(values_only=True))

            # Extract header info from first ~8 rows
            bank_name = ''
            company_name = ''
            account_number = ''
            currency = 'RUB'
            opening_balance = ''

            for i, row in enumerate(all_rows[:8]):
                cell_a = str(row[0] or '') if row and row[0] else ''
                if not cell_a:
                    continue
                # Row 1 (index 1): Bank name (e.g., "ООО РНКО "ЦИФРОВЫЕ РЕШЕНИЯ" г. Москва")
                # Match bank by specific patterns (РНКО, starts with "Банк", or contains " банк " as separate word)
                is_bank = ('РНКО' in cell_a or cell_a.startswith('Банк') or
                           re.search(r'(?:^|\s)банк\b|\bбанк(?:\s|$)', cell_a.lower()) is not None)
                if is_bank and not bank_name:
                    bank_name = cell_a.strip()
                # Row 2 (index 2): Company name — if bank already found, this is the company
                elif i == 2 and bank_name:
                    company_name = cell_a.strip()
                # If bank not yet found, company comes first, then bank
                elif ('ООО' in cell_a or 'ОсОО' in cell_a or 'АО' in cell_a or 'ИП' in cell_a) and not company_name:
                    company_name = cell_a.strip()
                # Row 4: Account number (e.g., "Выписка по счету N 40807810800000000086")
                if 'счет' in cell_a.lower():
                    m = re.search(r'(\d{20})', cell_a)
                    if m:
                        account_number = m.group(1)
                # Row 5: Currency
                if 'валют' in cell_a.lower():
                    if 'RUB' in cell_a: currency = 'RUB'
                    elif 'USD' in cell_a: currency = 'USD'
                    elif 'EUR' in cell_a: currency = 'EUR'
                # Row 6: Opening balance
                if 'остаток' in cell_a.lower():
                    m = re.search(r'([\d.]+)', cell_a)
                    if m:
                        opening_balance = m.group(1)

            # Find header row (contains "Дебет") and sub-header row (contains column details)
            header_idx = -1
            debit_col = -1
            credit_col = -1
            doc_number_col = -1
            doc_date_col = -1
            purpose_col = -1
            # Sub-header columns for counterparty (under "Контрагент" group)
            cpty_inn_col = -1
            cpty_name_col = -1
            cpty_account_col = -1
            cpty_bik_col = -1
            cpty_corr_col = -1
            cpty_bank_col = -1

            # First pass: find the row containing "Дебет"
            for i, row in enumerate(all_rows):
                for j, cell in enumerate(row):
                    if str(cell or '').strip() == 'Дебет':
                        header_idx = i
                        debit_col = j
                        break
                if header_idx >= 0:
                    break

            # Second pass: find all columns in the header row
            if header_idx >= 0:
                hrow = all_rows[header_idx]
                for j, cell in enumerate(hrow):
                    cell_str = str(cell or '').strip()
                    if cell_str == 'Кредит':
                        credit_col = j
                    elif cell_str.startswith('Номер док') or cell_str == 'Номер докум.':
                        doc_number_col = j
                    elif cell_str.startswith('Дата док') or cell_str == 'Дата докум.':
                        doc_date_col = j
                    elif cell_str == 'Назначение платежа':
                        purpose_col = j

            # Parse sub-header row (header_idx + 1) for counterparty column positions
            if header_idx >= 0 and header_idx + 1 < len(all_rows):
                sub_row = all_rows[header_idx + 1]
                # The "Контрагент" group starts at col 7: ИНН, КПП, Наименование, Счет, БИК, Коррсчет, Банк
                for j, cell in enumerate(sub_row or []):
                    cell_str = str(cell or '').strip()
                    if cell_str == 'ИНН' and cpty_inn_col < 0:
                        cpty_inn_col = j
                    elif cell_str == 'Наименование' and cpty_name_col < 0:
                        cpty_name_col = j
                    elif cell_str == 'Счет' and cpty_account_col < 0:
                        cpty_account_col = j
                    elif cell_str == 'БИК' and cpty_bik_col < 0:
                        cpty_bik_col = j
                    elif cell_str == 'Коррсчет' and cpty_corr_col < 0:
                        cpty_corr_col = j
                    elif cell_str == 'Банк' and cpty_bank_col < 0:
                        cpty_bank_col = j

            logger.info(f"XLSX header at row {header_idx + 1}, Дебет=col {debit_col}, Кредит=col {credit_col}, "
                        f"Контрагент: ИНН={cpty_inn_col}, Наим={cpty_name_col}, Счет={cpty_account_col}, "
                        f"БИК={cpty_bik_col}, Коррсчет={cpty_corr_col}, Банк={cpty_bank_col}")

            # Data starts 2 rows after header (header + sub-header)
            data_start = header_idx + 2 if header_idx >= 0 else 10
            data_rows = all_rows[data_start:]

            def safe_cell(row, col):
                """Get cell value safely, return empty string if not available."""
                if col < 0 or col >= len(row):
                    return ''
                val = row[col]
                if val is None:
                    return ''
                if hasattr(val, 'strftime'):
                    return val.strftime('%d.%m.%Y')
                return str(val).strip()

            credit_total = 0.0
            credit_transactions = []
            dates = []

            for row in data_rows:
                if not row or all(c is None or c == '' for c in row[:6]):
                    continue
                first_cell = row[0] if row else None
                if first_cell and isinstance(first_cell, str):
                    break  # Summary section reached
                try:
                    credit_val = row[credit_col] if credit_col >= 0 and credit_col < len(row) else 0
                    credit = float(credit_val or 0)
                    # Collect dates
                    date_val = row[0] if row[0] else None
                    if date_val:
                        if hasattr(date_val, 'strftime'):
                            dates.append(date_val.strftime('%d.%m.%Y'))
                        else:
                            dates.append(str(date_val))
                    # Only credit > 0
                    if credit > 0:
                        credit_total += credit
                        txn = {
                            'bik': safe_cell(row, cpty_bik_col),
                            'corr_account': safe_cell(row, cpty_corr_col),
                            'bank_name': safe_cell(row, cpty_bank_col),
                            'counterparty_account': safe_cell(row, cpty_account_col),
                            'counterparty': safe_cell(row, cpty_name_col),
                            'doc_number': safe_cell(row, doc_number_col),
                            'doc_date': safe_cell(row, doc_date_col),
                            'credit': f'{credit:,.2f}',
                            'purpose': safe_cell(row, purpose_col),
                            'counterparty_inn': safe_cell(row, cpty_inn_col),
                        }
                        credit_transactions.append(txn)
                except (ValueError, TypeError, IndexError):
                    pass

            period_start = min(dates) if dates else ''
            period_end = max(dates) if dates else ''

            # Build field results
            f['account_number'] = FieldResult(account_number, 1.0, 'bank_statement_parsing')
            f['account_holder'] = FieldResult(company_name, 1.0, 'bank_statement_parsing')
            f['period_start'] = FieldResult(period_start, 1.0, 'bank_statement_parsing')
            f['period_end'] = FieldResult(period_end, 1.0, 'bank_statement_parsing')
            f['total_credit'] = FieldResult(f'{credit_total:,.2f} ₽', 1.0, 'bank_statement_parsing')
            f['transaction_count'] = FieldResult(str(len(credit_transactions)), 1.0, 'bank_statement_parsing')
            f['currency'] = FieldResult(currency, 1.0, 'bank_statement_parsing')
            f['transactions'] = FieldResult(credit_transactions, 1.0, 'bank_statement_parsing')

            return f

        except Exception as e:
            logger.error(f"XLSX parsing error: {e}")
            return f

    def _validate(self, doc_type: str, fields: dict) -> dict:
        """Basic validation."""
        v = {}
        if doc_type == "snils":
            snils = fields.get("snils_number", {}).get("value", "")
            digits = re.sub(r"\D", "", snils)
            if len(digits) == 11:
                num = digits[:9]
                ctrl = int(digits[9:11])
                total = sum(int(d) * (9 - i) for i, d in enumerate(num))
                if total < 100:
                    exp = total
                elif total in (100, 101):
                    exp = 0
                else:
                    r = total % 101
                    exp = 0 if r >= 100 else r
                v["snils_number"] = {"valid": ctrl == exp, "message": f"Checksum {'OK' if ctrl == exp else 'INVALID'}"}

        if doc_type == "inn":
            inn = fields.get("inn_number", {}).get("value", "")
            digits = re.sub(r"\D", "", inn)
            if len(digits) == 12:
                d = [int(c) for c in digits]
                w11 = [7,2,4,10,3,5,9,4,6,8]
                w12 = [3,7,2,4,10,3,5,9,4,6,8]
                c11 = sum(w*d[i] for i,w in enumerate(w11)) % 11 % 10
                c12 = sum(w*d[i] for i,w in enumerate(w12)) % 11 % 10
                ok = d[10] == c11 and d[11] == c12
                v["inn_number"] = {"valid": ok, "message": f"Checksum {'OK' if ok else 'INVALID'}"}

        return v

    def _demo_result(self, result: dict, start: float) -> dict:
        """Return a demo result when OCR is not available."""
        result["document_type"] = "passport_rf"
        result["classification_confidence"] = 0.95
        result["overall_confidence"] = 0.92
        result["fields"] = {
            "last_name": {"value": "ИВАНОВ", "confidence": 0.98, "source": "demo", "auto_fill": True, "needs_review": False},
            "first_name": {"value": "ИВАН", "confidence": 0.97, "source": "demo", "auto_fill": True, "needs_review": False},
            "patronymic": {"value": "ИВАНОВИЧ", "confidence": 0.96, "source": "demo", "auto_fill": True, "needs_review": False},
            "birth_date": {"value": "01.01.1990", "confidence": 0.99, "source": "demo", "auto_fill": True, "needs_review": False},
            "series": {"value": "45 20", "confidence": 0.95, "source": "demo", "auto_fill": True, "needs_review": False},
            "number": {"value": "123456", "confidence": 0.88, "source": "demo", "auto_fill": False, "needs_review": True},
            "sex": {"value": "М", "confidence": 0.92, "source": "demo", "auto_fill": False, "needs_review": True},
        }
        result["warnings"].append("DEMO MODE: No OCR engines available, showing sample data")
        result["processing_time_ms"] = int((time.time() - start) * 1000)
        return result


pipeline = LightweightPipeline()


# ============================================================
# Custom JSON encoder for numpy types
# ============================================================
import json as _json

class NumpyEncoder(_json.JSONEncoder):
    """JSON encoder that handles numpy types."""
    def default(self, obj):
        if HAS_CV2:
            import numpy as _np
            if isinstance(obj, (_np.integer,)):
                return int(obj)
            if isinstance(obj, (_np.floating,)):
                return float(obj)
            if isinstance(obj, (_np.bool_,)):
                return bool(obj)
            if isinstance(obj, (_np.ndarray,)):
                return obj.tolist()
        return super().default(obj)

def sanitize_for_json(obj):
    """Recursively convert numpy types to Python native types."""
    if HAS_CV2:
        import numpy as _np
        if isinstance(obj, _np.bool_):
            return bool(obj)
        if isinstance(obj, _np.integer):
            return int(obj)
        if isinstance(obj, _np.floating):
            return float(obj)
        if isinstance(obj, _np.ndarray):
            return obj.tolist()
    if isinstance(obj, dict):
        return {k: sanitize_for_json(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [sanitize_for_json(i) for i in obj]
    return obj

# ============================================================
# FastAPI App
# ============================================================
app = FastAPI(
    title="DocLens API",
    description="Document Recognition SaaS",
    version="1.0.0",
    docs_url="/api/docs",
    redoc_url="/api/redoc",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Try to mount static/templates
BASE_DIR = Path(__file__).parent
if (BASE_DIR / "static").exists():
    app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")


# ============================================================
# Auth dependency
# ============================================================
def get_tenant(request: Request) -> dict:
    api_key = request.headers.get("X-API-Key", "")
    if not api_key:
        raise HTTPException(401, "API key required (X-API-Key header)")
    tenant = db.auth_by_key(api_key)
    if not tenant:
        raise HTTPException(401, "Invalid API key")
    return tenant


# ============================================================
# Web UI (inline HTML if templates not available)
# ============================================================

LANDING_HTML = """<!DOCTYPE html>
<html lang="ru"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>DocLens — Document Recognition SaaS</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}body{font-family:system-ui,sans-serif;background:#f8fafc;color:#1e293b}
.hero{background:linear-gradient(135deg,#1e40af,#3b82f6);color:white;padding:60px 20px;text-align:center}
.hero h1{font-size:36px;margin-bottom:12px}.hero p{font-size:18px;opacity:.9;max-width:600px;margin:0 auto 24px}
.btn{display:inline-block;padding:12px 28px;background:white;color:#1e40af;border-radius:8px;text-decoration:none;font-weight:600;margin:6px}
.btn-outline{background:transparent;border:2px solid white;color:white}
.container{max-width:1000px;margin:0 auto;padding:40px 20px}
.grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:16px;margin:24px 0}
.card{background:white;border-radius:10px;padding:20px;border:1px solid #e2e8f0;text-align:center}
.card h3{font-size:15px;margin-bottom:8px}.card p{font-size:13px;color:#64748b}
.card .icon{font-size:28px;margin-bottom:8px}
h2{font-size:22px;text-align:center;margin-bottom:20px}
pre{background:#1e293b;color:#e2e8f0;padding:20px;border-radius:10px;overflow-x:auto;font-size:13px;margin:16px 0}
.pricing{display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:16px}
.price-card{background:white;border-radius:10px;padding:24px;border:1px solid #e2e8f0;text-align:center}
.price-card.featured{border-color:#3b82f6;box-shadow:0 0 0 2px #3b82f6}
.price-card h3{font-size:18px}.price-card .price{font-size:28px;font-weight:700;color:#1e40af;margin:8px 0}
.price-card .price small{font-size:14px;color:#94a3b8;font-weight:400}
.price-card ul{list-style:none;text-align:left;font-size:13px;margin:12px 0}
.price-card li{padding:4px 0;color:#475569}
.price-card li::before{content:"✓ ";color:#22c55e;font-weight:bold}
footer{text-align:center;padding:30px;color:#94a3b8;font-size:13px}
footer a{color:#3b82f6;text-decoration:none}
</style></head><body>
<div class="hero">
<h1>DocLens</h1>
<p>Распознавание документов как сервис. Извлекайте данные из паспортов, СНИЛС, ИНН и водительских удостоверений через простой API.</p>
<a href="/demo" class="btn">Попробовать демо</a>
<a href="/api/docs" class="btn btn-outline">API документация</a>
</div>
<div class="container">
<h2>Поддерживаемые документы</h2>
<div class="grid">
<div class="card"><div class="icon">🪪</div><h3>Паспорт РФ</h3><p>Серия, номер, ФИО, дата рождения, место, код подразделения</p></div>
<div class="card"><div class="icon">🌍</div><h3>Паспорта СНГ</h3><p>MRZ-парсинг паспортов 12 стран СНГ (Узбекистан, Казахстан...)</p></div>
<div class="card"><div class="icon">🚗</div><h3>Водительское удостоверение</h3><p>Номер, ФИО, категории, даты выдачи и действия</p></div>
<div class="card"><div class="icon">📋</div><h3>СНИЛС</h3><p>Номер с валидацией контрольной суммы, ФИО, дата рождения</p></div>
<div class="card"><div class="icon">📄</div><h3>ИНН</h3><p>Номер (10/12 цифр) с проверкой контрольных цифр</p></div>
</div>

<h2 style="margin-top:40px">Быстрый старт</h2>
<pre>
# 1. Зарегистрируйтесь
curl -X POST http://HOST/api/v1/tenants \\
  -H "Content-Type: application/json" \\
  -d '{"name": "My Company", "email": "dev@company.com"}'

# 2. Распознайте документ
curl -X POST http://HOST/api/v1/recognize \\
  -H "X-API-Key: dl_live_your_key_here" \\
  -F "file=@passport.jpg"
</pre>

<h2 style="margin-top:40px">Тарифы</h2>
<div class="pricing">
<div class="price-card"><h3>Free</h3><div class="price">0 ₽<small>/мес</small></div><ul><li>100 запросов/день</li><li>5 типов документов</li><li>REST API</li></ul></div>
<div class="price-card featured"><h3>Basic</h3><div class="price">9 990 ₽<small>/мес</small></div><ul><li>1 000 запросов/день</li><li>Приоритетная обработка</li><li>Webhook уведомления</li></ul></div>
<div class="price-card"><h3>Pro</h3><div class="price">29 990 ₽<small>/мес</small></div><ul><li>10 000 запросов/день</li><li>Проверка по базам ФМС/ФНС</li><li>Выделенная поддержка</li></ul></div>
<div class="price-card"><h3>Enterprise</h3><div class="price">По запросу</div><ul><li>100 000+ запросов/день</li><li>On-premise установка</li><li>SLA 99.9%</li><li>Кастомные парсеры</li></ul></div>
</div>
</div>
<footer>DocLens v1.0.0 &bull; <a href="/demo">Демо</a> &bull; <a href="/dashboard">Дашборд</a> &bull; <a href="/api/docs">API Docs</a></footer>
</body></html>"""

DEMO_HTML = """<!DOCTYPE html>
<html lang="ru"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>DocLens — Демо</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}body{font-family:system-ui,sans-serif;background:#f1f5f9;color:#1e293b}
.header{background:#1e40af;color:white;padding:16px 24px;display:flex;align-items:center;justify-content:space-between}
.header h1{font-size:20px}.header a{color:white;text-decoration:none;opacity:.8}
.container{max-width:900px;margin:24px auto;padding:0 20px}
.card{background:white;border-radius:12px;padding:24px;margin-bottom:20px;border:1px solid #e2e8f0}
h2{font-size:18px;margin-bottom:14px}
label{display:block;font-size:13px;font-weight:600;margin-bottom:6px;color:#475569}
input,select{width:100%;padding:10px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;margin-bottom:12px}
.upload-area{border:2px dashed #93c5fd;border-radius:10px;padding:40px;text-align:center;cursor:pointer;transition:.2s}
.upload-area:hover{background:#eff6ff;border-color:#3b82f6}
.upload-area.active{border-color:#22c55e;background:#f0fdf4}
.btn{display:inline-block;padding:12px 28px;background:#2563eb;color:white;border:none;border-radius:8px;font-size:15px;font-weight:600;cursor:pointer;width:100%}
.btn:hover{background:#1d4ed8}.btn:disabled{background:#94a3b8;cursor:not-allowed}
.result{display:none}.result.show{display:block}
.field-row{display:flex;align-items:center;gap:12px;padding:8px 0;border-bottom:1px solid #f1f5f9}
.field-name{width:140px;font-weight:600;font-size:13px;color:#64748b}
.field-value{flex:1;font-size:15px}
.conf-bar{width:80px;height:8px;background:#e2e8f0;border-radius:4px;overflow:hidden}
.conf-fill{height:100%;border-radius:4px}
.badge{display:inline-block;padding:2px 8px;border-radius:12px;font-size:11px;font-weight:600}
.badge-green{background:#dcfce7;color:#166534}.badge-yellow{background:#fef9c3;color:#854d0e}.badge-red{background:#fee2e2;color:#991b1b}
.badge-blue{background:#dbeafe;color:#1e40af}
.spinner{display:none;text-align:center;padding:20px}.spinner.show{display:block}
.warn{background:#fffbeb;border:1px solid #fde68a;border-radius:8px;padding:10px;margin-top:10px;font-size:13px;color:#92400e}
</style></head><body>
<div class="header"><h1>DocLens Demo</h1><a href="/">← На главную</a></div>
<div class="container">
<div class="card">
<h2>1. API-ключ</h2>
<label>Введите ваш API-ключ (или зарегистрируйтесь ниже)</label>
<input type="text" id="apiKey" placeholder="dl_live_...">
<button onclick="register()" style="background:#059669;color:white;border:none;padding:8px 16px;border-radius:6px;cursor:pointer;font-size:13px">Получить бесплатный ключ</button>
<div id="regResult" style="margin-top:8px;font-size:13px;color:#059669"></div>
</div>
<div class="card">
<h2>2. Загрузите документ</h2>
<div class="upload-area" id="dropZone" onclick="document.getElementById('fileInput').click()">
<div style="font-size:32px;margin-bottom:8px">📄</div>
<p>Перетащите файл сюда или кликните для выбора</p>
<p style="font-size:12px;color:#94a3b8;margin-top:4px">JPEG, PNG, PDF — до 20MB</p>
</div>
<input type="file" id="fileInput" accept="image/*,.pdf" style="display:none">
<div id="fileName" style="margin-top:8px;font-size:13px;color:#64748b"></div>
</div>
<div class="card">
<h2>3. Тип документа</h2>
<select id="docType">
<option value="">Автоопределение</option>
<option value="passport_rf">Паспорт РФ</option>
<option value="passport_cis">Паспорт СНГ</option>
<option value="driver_license">Водительское удостоверение</option>
<option value="snils">СНИЛС</option>
<option value="inn">ИНН</option>
</select>
<button class="btn" id="recognizeBtn" onclick="recognize()" disabled>Распознать</button>
</div>
<div class="spinner" id="spinner"><p>Обработка документа...</p></div>
<div class="result card" id="resultCard">
<h2>Результат</h2>
<div id="resultMeta" style="margin-bottom:16px"></div>
<div id="resultFields"></div>
<div id="resultValidation"></div>
<div id="resultWarnings"></div>
</div>
</div>
<script>
let selectedFile=null;
const dropZone=document.getElementById('dropZone'),fileInput=document.getElementById('fileInput');
dropZone.addEventListener('dragover',e=>{e.preventDefault();dropZone.classList.add('active')});
dropZone.addEventListener('dragleave',()=>dropZone.classList.remove('active'));
dropZone.addEventListener('drop',e=>{e.preventDefault();dropZone.classList.remove('active');if(e.dataTransfer.files.length){selectFile(e.dataTransfer.files[0])}});
fileInput.addEventListener('change',()=>{if(fileInput.files.length)selectFile(fileInput.files[0])});
function selectFile(f){selectedFile=f;document.getElementById('fileName').textContent='Выбран: '+f.name+' ('+Math.round(f.size/1024)+'KB)';document.getElementById('recognizeBtn').disabled=false;dropZone.classList.add('active')}
async function register(){
try{const r=await fetch('/api/v1/tenants',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({name:'Demo User',email:'demo_'+Date.now()+'@test.com'})});
const d=await r.json();if(r.ok){document.getElementById('apiKey').value=d.api_key;document.getElementById('regResult').textContent='Ключ создан! Сохраните его: '+d.api_key}else{document.getElementById('regResult').textContent='Ошибка: '+(d.detail||JSON.stringify(d));document.getElementById('regResult').style.color='red'}}
catch(e){document.getElementById('regResult').textContent='Ошибка: '+e.message}}
async function recognize(){
const key=document.getElementById('apiKey').value;if(!key){alert('Введите API-ключ');return}if(!selectedFile){alert('Выберите файл');return}
document.getElementById('spinner').classList.add('show');document.getElementById('resultCard').classList.remove('show');
const fd=new FormData();fd.append('file',selectedFile);
const dt=document.getElementById('docType').value;if(dt)fd.append('document_type',dt);
try{const r=await fetch('/api/v1/recognize',{method:'POST',headers:{'X-API-Key':key},body:fd});const d=await r.json();
document.getElementById('spinner').classList.remove('show');
if(r.ok){showResult(d)}else{alert('Ошибка: '+(d.detail||JSON.stringify(d)))}}
catch(e){document.getElementById('spinner').classList.remove('show');alert('Ошибка: '+e.message)}}
function showResult(d){
const rc=document.getElementById('resultCard');rc.classList.add('show');
const typeNames={passport_rf:'Паспорт РФ',passport_cis:'Паспорт СНГ',driver_license:'ВУ',snils:'СНИЛС',inn:'ИНН',unknown:'Неизвестный'};
document.getElementById('resultMeta').innerHTML=
'<span class="badge badge-blue">'+( typeNames[d.document_type]||d.document_type)+'</span> '+
'<span class="badge '+(d.overall_confidence>=0.9?'badge-green':d.overall_confidence>=0.7?'badge-yellow':'badge-red')+'">'+
Math.round(d.overall_confidence*100)+'% confidence</span> '+
'<span style="font-size:12px;color:#94a3b8;margin-left:8px">'+d.processing_time_ms+'ms</span>';
const fieldNames={last_name:'Фамилия',first_name:'Имя',patronymic:'Отчество',birth_date:'Дата рождения',sex:'Пол',number:'Номер',series:'Серия',expiry_date:'Срок действия',issue_date:'Дата выдачи',issuer:'Кем выдан',department_code:'Код подразделения',birth_place:'Место рождения',residence:'Место жительства',country_code:'Код страны',snils_number:'СНИЛС',inn_number:'ИНН',inn_type:'Тип ИНН',categories:'Категории'};
let fhtml='';
for(const[name,f]of Object.entries(d.fields||{})){
const pct=Math.round(f.confidence*100);
const color=pct>=95?'#22c55e':pct>=70?'#eab308':'#ef4444';
const label=fieldNames[name]||name;
fhtml+='<div class="field-row"><div class="field-name">'+label+'</div><div class="field-value">'+f.value+'</div>'+
'<div class="conf-bar"><div class="conf-fill" style="width:'+pct+'%;background:'+color+'"></div></div>'+
'<span style="font-size:12px;color:#94a3b8;width:35px">'+pct+'%</span>'+
(f.auto_fill?'<span class="badge badge-green">auto</span>':f.needs_review?'<span class="badge badge-yellow">review</span>':'')+
'</div>'}
document.getElementById('resultFields').innerHTML=fhtml;
let vhtml='';for(const[name,v]of Object.entries(d.validation||{})){
vhtml+='<div style="font-size:13px;margin:4px 0"><span class="badge '+(v.valid?'badge-green':'badge-red')+'">'+(v.valid?'✓':'✗')+'</span> '+name+': '+v.message+'</div>'}
document.getElementById('resultValidation').innerHTML=vhtml;
let whtml='';if(d.warnings&&d.warnings.length){whtml='<div class="warn"><b>Предупреждения:</b><br>'+d.warnings.join('<br>')+'</div>'}
document.getElementById('resultWarnings').innerHTML=whtml}
</script></body></html>"""


@app.get("/webapp", response_class=HTMLResponse)
async def telegram_webapp():
    """Serve Telegram MiniApp HTML."""
    webapp_path = BASE_DIR / "telegram_bot" / "webapp.html"
    if webapp_path.exists():
        return HTMLResponse(webapp_path.read_text(encoding="utf-8"))
    return HTMLResponse("<h1>MiniApp not found</h1>", status_code=404)


@app.get("/", response_class=HTMLResponse)
async def landing():
    tpl = BASE_DIR / "templates" / "index.html"
    if tpl.exists():
        from jinja2 import Template
        return Template(tpl.read_text()).render(settings=type("S", (), {"APP_VERSION": "1.0.0"})())
    return HTMLResponse(LANDING_HTML)


@app.get("/demo", response_class=HTMLResponse)
async def demo():
    demo_key = getattr(app.state, "demo_api_key", "")
    # Try to serve from static/demo.html first (production version)
    static_demo = BASE_DIR / "static" / "demo.html"
    if static_demo.exists():
        html = static_demo.read_text()
        # Inject demo key script in <head> BEFORE main script so it's available when JS runs
        inject = f'\n<script>window.__demoApiKey="{demo_key}";</script>\n'
        html = html.replace("</head>", inject + "</head>")
        return HTMLResponse(html)
    # Fallback to templates/demo.html
    tpl = BASE_DIR / "templates" / "demo.html"
    if tpl.exists():
        from jinja2 import Template
        html = tpl.read_text()
        # Inject demo key script in <head> BEFORE main script
        inject = f'\n<script>window.__demoApiKey="{demo_key}";</script>\n'
        html = html.replace("</head>", inject + "</head>")
        return HTMLResponse(Template(html).render(settings=type("S", (), {"APP_VERSION": "1.0.0"})()))
    # Fallback to inline HTML (for backwards compatibility)
    html = DEMO_HTML.replace('placeholder="dl_live_..."', f'placeholder="dl_live_..." value="{demo_key}"')
    return HTMLResponse(html)


# ============================================================
# API Routes
# ============================================================

@app.get("/api/v1/health")
async def health():
    engine = "easyocr" if HAS_EASYOCR else ("tesseract" if HAS_TESSERACT else "none")
    has_rus = False
    tess_langs = []
    if HAS_TESSERACT:
        try:
            tess_langs = pytesseract.get_languages()
            has_rus = "rus" in tess_langs
        except Exception:
            pass
    return {
        "status": "ok", "version": "1.0.0",
        "ocr_engine": engine,
        "ocr_ready": HAS_EASYOCR or HAS_TESSERACT,
        "opencv": HAS_CV2,
        "tesseract_languages": tess_langs,
        "has_russian": has_rus,
        "warning": None if (HAS_EASYOCR or has_rus) else "⚠️ No OCR engines with Russian support! Install EasyOCR or Tesseract with Russian language: brew install tesseract-lang"
    }


@app.post("/api/v1/tenants")
async def register_tenant(request: Request):
    data = await request.json()
    name = data.get("name", "")
    email = data.get("email", "")
    if not name or not email:
        raise HTTPException(400, "name and email required")

    # Check duplicate
    for t in db.tenants.values():
        if t["email"] == email:
            raise HTTPException(409, "Email already registered")

    tenant = db.create_tenant(name, email)
    key_record, raw_key = db.create_api_key(tenant["id"])

    return {
        "id": key_record["id"],
        "name": key_record["name"],
        "key_prefix": key_record["key_prefix"],
        "created_at": key_record["created_at"],
        "api_key": raw_key,
        "tenant_id": tenant["id"],
    }


@app.get("/api/v1/tenants/me")
async def tenant_info(tenant: dict = Depends(get_tenant)):
    return tenant


@app.post("/api/v1/recognize")
async def recognize(
    request: Request,
    file: UploadFile = File(...),
    document_type: str = Form(default=None),
    tenant: dict = Depends(get_tenant),
):
    """Recognize a document."""
    allowed = ["image/jpeg", "image/png", "image/tiff", "image/heic", "image/heif", "application/pdf", "application/octet-stream",
               "text/plain", "text/csv", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
               "application/vnd.ms-excel", "application/csv"]
    # Also allow by file extension for HEIC and other types where content_type may be generic
    ext = (file.filename or "").lower().rsplit(".", 1)[-1] if file.filename else ""
    allowed_ext = {"jpg", "jpeg", "png", "tiff", "tif", "heic", "heif", "pdf", "bmp", "webp", "txt", "csv", "xlsx", "xls"}
    if file.content_type and file.content_type not in allowed and ext not in allowed_ext:
        raise HTTPException(400, f"Unsupported file type: {file.content_type}")

    image_bytes = await file.read()
    if len(image_bytes) > 20 * 1024 * 1024:
        raise HTTPException(400, "File too large (max 20MB)")

    logger.info(f"Processing {file.filename} ({len(image_bytes)} bytes) for tenant {tenant['id']}")

    result = pipeline.process(image_bytes, document_type_hint=document_type)
    result["original_filename"] = file.filename

    # Apply ML corrections to extracted fields
    if result.get("fields") and result.get("document_type", "unknown") != "unknown":
        try:
            result["fields"] = ml_correct_fields(result["document_type"], result["fields"])
            logger.info("ML corrections applied")
        except Exception as e:
            logger.warning(f"ML correction failed (non-fatal): {e}")

    # Sanitize numpy types for JSON serialization
    result = sanitize_for_json(result)

    rec = db.save_recognition(tenant["id"], result)
    return {**result, "id": rec["id"]}


@app.get("/api/v1/recognitions")
async def list_recognitions(tenant: dict = Depends(get_tenant)):
    recs = [r for r in db.recognitions.values() if r["tenant_id"] == tenant["id"]]
    recs.sort(key=lambda r: r["created_at"], reverse=True)
    return recs[:50]


@app.get("/api/v1/recognitions/{rid}")
async def get_recognition(rid: str, tenant: dict = Depends(get_tenant)):
    rec = db.recognitions.get(rid)
    if not rec or rec["tenant_id"] != tenant["id"]:
        raise HTTPException(404, "Not found")
    return rec


@app.get("/api/v1/usage")
async def get_usage(tenant: dict = Depends(get_tenant)):
    return db.get_usage(tenant["id"])


@app.post("/api/v1/api-keys")
async def create_key(request: Request, tenant: dict = Depends(get_tenant)):
    data = await request.json() if request.headers.get("content-type", "").startswith("application/json") else {}
    name = data.get("name", "New Key")
    record, raw_key = db.create_api_key(tenant["id"], name)
    return {**record, "api_key": raw_key}


@app.get("/api/v1/api-keys")
async def list_keys(tenant: dict = Depends(get_tenant)):
    return [k for k in db.api_keys.values() if k["tenant_id"] == tenant["id"]]


@app.post("/api/v1/debug-ocr")
async def debug_ocr(
    request: Request,
    file: UploadFile = File(...),
    tenant: dict = Depends(get_tenant),
):
    """Debug endpoint: returns raw OCR text and MRZ analysis."""
    image_bytes = await file.read()
    debug_info = {"engine": pipeline.engine, "raw_lines": [], "mrz_candidates": [], "dedicated_mrz": []}

    if HAS_CV2:
        nparr = np.frombuffer(image_bytes, np.uint8)
        image = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        if image is None:
            return {"error": "Cannot decode image"}
        h, w = image.shape[:2]
        debug_info["image_size"] = f"{w}x{h}"
        if max(h, w) > 2000:
            scale = 2000 / max(h, w)
            image = cv2.resize(image, (int(w * scale), int(h * scale)))
            h, w = image.shape[:2]

        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
        enhanced = clahe.apply(gray)
        image_enhanced = cv2.cvtColor(enhanced, cv2.COLOR_GRAY2BGR)

        if HAS_TESSERACT:
            ocr_lines = pipeline._ocr_tesseract(image_enhanced)
            debug_info["raw_lines"] = [{"text": l.text, "confidence": float(round(l.confidence, 3))} for l in ocr_lines]

            # Also run dedicated MRZ pass
            try:
                mrz_lines = pipeline._ocr_mrz_zone(image)
                debug_info["dedicated_mrz"] = mrz_lines
            except Exception as e:
                debug_info["dedicated_mrz_error"] = str(e)

            # Check tess languages
            try:
                debug_info["tess_languages"] = pytesseract.get_languages()
            except Exception:
                debug_info["tess_languages"] = ["unknown"]

    return debug_info


# ============================================================
# ML / Active Learning API
# ============================================================

@app.post("/api/v1/ml/corrections")
async def submit_corrections(request: Request, tenant: dict = Depends(get_tenant)):
    """Submit field corrections for ML learning.

    Body: {
        "recognition_id": "...",
        "document_type": "passport_rf",
        "corrections": [
            {"field_name": "last_name", "original_value": "ЛЪЯКОВ", "corrected_value": "ДЬЯКОВ", "confidence": 0.8}
        ]
    }
    """
    data = await request.json()
    recognition_id = data.get("recognition_id", "")
    document_type = data.get("document_type", "")
    corrections_data = data.get("corrections", [])

    if not corrections_data:
        raise HTTPException(400, "No corrections provided")

    saved = []
    for c in corrections_data:
        result = db.save_correction(
            tenant_id=tenant["id"],
            recognition_id=recognition_id,
            document_type=document_type,
            field_name=c.get("field_name", ""),
            original_value=c.get("original_value", ""),
            corrected_value=c.get("corrected_value", ""),
            confidence=c.get("confidence", 0.0),
        )
        if result:
            saved.append(result)

    return {
        "status": "ok",
        "corrections_saved": len(saved),
        "message": f"Сохранено {len(saved)} коррекций для обучения",
    }


@app.get("/api/v1/ml/corrections")
async def list_corrections(
    document_type: str = None,
    tenant: dict = Depends(get_tenant),
):
    """List corrections submitted by this tenant."""
    corrections = [
        c for c in db.corrections.values()
        if c["tenant_id"] == tenant["id"]
        and (document_type is None or c["document_type"] == document_type)
    ]
    corrections.sort(key=lambda c: c["created_at"], reverse=True)
    return corrections[:100]


@app.get("/api/v1/ml/corrections/stats")
async def correction_stats(
    document_type: str = None,
    tenant: dict = Depends(get_tenant),
):
    """Get correction and pattern statistics."""
    return db.get_correction_stats(document_type)


@app.get("/api/v1/ml/patterns")
async def list_patterns(
    document_type: str = None,
    tenant: dict = Depends(get_tenant),
):
    """List learned correction patterns."""
    patterns = [
        p for p in db.correction_patterns.values()
        if p["is_active"]
        and (document_type is None or p["document_type"] == document_type)
    ]
    patterns.sort(key=lambda p: p["occurrence_count"], reverse=True)
    return patterns


@app.post("/api/v1/ml/train")
async def train_model(request: Request, tenant: dict = Depends(get_tenant)):
    """Train ML model from accumulated corrections.

    Body (optional): {"document_type": "passport_rf", "force": false}
    """
    try:
        data = await request.json()
    except Exception:
        data = {}

    document_type = data.get("document_type")
    force = data.get("force", False)

    result = ml_train_model(document_type, force)
    return result


@app.get("/api/v1/ml/train/status")
async def training_status(tenant: dict = Depends(get_tenant)):
    """Get ML system status."""
    last_run = db.training_runs[-1] if db.training_runs else None
    pending = sum(1 for c in db.corrections.values() if c["status"] in ("pending", "approved"))

    return {
        "last_training_run": last_run,
        "active_patterns": len([p for p in db.correction_patterns.values() if p["is_active"]]),
        "pending_corrections": pending,
        "ml_model_loaded": db.ml_model_data is not None,
        "ready_for_training": pending >= 20,
    }


@app.post("/api/v1/ml/correct")
async def test_correction(request: Request, tenant: dict = Depends(get_tenant)):
    """Test ML correction on arbitrary fields (debug endpoint).

    Body: {"document_type": "passport_rf", "fields": {"last_name": {"value": "ЛЪЯКОВ", "confidence": 0.7}}}
    """
    data = await request.json()
    document_type = data.get("document_type", "")
    fields = data.get("fields", {})

    corrected = ml_correct_fields(document_type, fields)
    return {
        "document_type": document_type,
        "original": fields,
        "corrected": corrected,
    }


# ============================================================
# Run
# ============================================================
if __name__ == "__main__":
    logger.info("Starting DocLens (standalone mode)...")
    logger.info(f"  OpenCV: {'✓' if HAS_CV2 else '✗'}")
    logger.info(f"  EasyOCR: {'✓' if HAS_EASYOCR else '✗'}")
    logger.info(f"  Tesseract: {'✓' if HAS_TESSERACT else '✗'}")

    # Create demo tenant + key automatically
    demo_tenant = db.create_tenant("Demo", "demo@doclens.local")
    demo_key_rec, demo_raw_key = db.create_api_key(demo_tenant["id"], "demo-key")
    logger.info(f"  Demo API key: {demo_raw_key}")
    # Store demo key globally so templates can use it
    app.state.demo_api_key = demo_raw_key

    port = int(os.environ.get("PORT", 7860))
    uvicorn.run(app, host="0.0.0.0", port=port, log_level="info")
